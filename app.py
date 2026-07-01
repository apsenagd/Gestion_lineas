from math import ceil
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response, abort, g, session
from werkzeug.security import check_password_hash, generate_password_hash
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
try:
    from passlib.hash import scrypt as passlib_scrypt
    PASSLIB_SCRYPT_AVAILABLE = True
except Exception:
    PASSLIB_SCRYPT_AVAILABLE = False
import os
import mysql.connector
from jinja2 import Environment
import socket
import hashlib
import base64
import time
import traceback
from datetime import datetime
import io
import csv
from io import BytesIO
import smtplib
from email.message import EmailMessage
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
app.secret_key = 'mi_clave_secreta_2026'



# Habilitar extensión 'do' en Jinja2
app.jinja_env.add_extension('jinja2.ext.do')


@app.template_filter('col_currency')
def col_currency(value):
    """Format number as Colombian-style currency: thousands with '.' and decimal ','; drop decimals if .00."""
    try:
        if value is None:
            return ''
        v = float(value)
    except Exception:
        return str(value)
    s = f"{v:,.2f}"  # 12,000.00
    # swap separators: comma->TEMP, dot->comma, TEMP->dot => 12.000,00
    s = s.replace(',', 'X').replace('.', ',').replace('X', '.')
    # drop decimal part if ,00
    if s.endswith(',00'):
        s = s[:-3]
    return s

def conectar_db():
    try:
        db = mysql.connector.connect(
            host="localhost",
            user="root",
            password="Mysql2026*",   #clave
            database="gestion_lineas",
            charset="utf8mb4",
            use_unicode=True,
            autocommit=True,
            connection_timeout=5
        )
    except Exception as e:
        print(f"Error conectando a la base de datos: {e}")
        raise
    # Aumentar timeout de lock y usar isolation más seguro
    cur = db.cursor()
    try:
        cur.execute("SET innodb_lock_wait_timeout = 20")
        cur.execute("SET SESSION TRANSACTION ISOLATION LEVEL READ COMMITTED")
    finally:
        cur.close()
    return db


def has_fecha_modificacion():
    """Check (and cache) whether the `lineas` table has a `fecha_modificacion` column."""
    if app.config.get('HAS_FECHA_MOD') is not None:
        return app.config['HAS_FECHA_MOD']
    try:
        db = conectar_db()
        cur = db.cursor()
        # Determine database name from connection if available
        db_name = getattr(db, 'database', None) or 'gestion_lineas'
        cur.execute("SELECT COUNT(1) FROM information_schema.columns WHERE table_schema = %s AND table_name = 'lineas' AND column_name = 'fecha_modificacion'", (db_name,))
        exists = cur.fetchone()[0] > 0
    except Exception:
        exists = False
    try:
        cur.close()
        db.close()
    except Exception:
        pass
    app.config['HAS_FECHA_MOD'] = exists
    return exists


def has_lineas_touch():
    """Check (and cache) whether the `lineas_touch` helper table exists."""
    if app.config.get('HAS_LINEAS_TOUCH') is not None:
        return app.config['HAS_LINEAS_TOUCH']
    try:
        db = conectar_db()
        cur = db.cursor()
        db_name = getattr(db, 'database', None) or 'gestion_lineas'
        cur.execute("SELECT COUNT(1) FROM information_schema.tables WHERE table_schema = %s AND table_name = 'lineas_touch'", (db_name,))
        exists = cur.fetchone()[0] > 0
    except Exception:
        exists = False
    try:
        cur.close(); db.close()
    except Exception:
        pass
    app.config['HAS_LINEAS_TOUCH'] = exists
    return exists


def get_estado_id(cur, nombre_estado):
    """Return the ID of an estado_linea row for the given state name."""
    if not nombre_estado:
        return None
    try:
        cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", (nombre_estado.lower(),))
        row = cur.fetchone()
        if not row:
            return None
        return row['id_estado'] if isinstance(row, dict) else row[0]
    except Exception:
        return None


# In-memory touch cache: maps id_linea -> datetime (UTC)
def touch_linea_in_memory(id_linea):
    try:
        cache = app.config.setdefault('LINEAS_TOUCH_CACHE', {})
        cache[int(id_linea)] = datetime.utcnow()
    except Exception:
        app.logger.exception('touch_linea_in_memory failed')


def get_touch_ts(id_linea):
    try:
        cache = app.config.get('LINEAS_TOUCH_CACHE', {})
        return cache.get(int(id_linea))
    except Exception:
        return None


def touch_lineas_for_usuario(db, id_usuario):
    try:
        cur = db.cursor()
        cur.execute("SELECT id_linea FROM lineas WHERE id_usuario = %s", (id_usuario,))
        rows = cur.fetchall()
        for r in rows:
            # fetch returns tuples for non-dict cursor
            lid = r[0] if isinstance(r, (list, tuple)) else r.get('id_linea')
            try:
                touch_linea_in_memory(lid)
            except Exception:
                pass
        try:
            cur.close()
        except Exception:
            pass
    except Exception:
        app.logger.exception('touch_lineas_for_usuario failed')


# Helper: execute with simple timeout wrapper (mysql-connector doesn't support per-query timeout param)
def execute_with_timeout(cur, query, params=None, timeout=2.0):
    start = time.time()
    cur.execute(query, params or [])
    duration = time.time() - start
    return duration

# =========================
# FUNCIÓN AUXILIAR: Registrar Novedad
# =========================
def registrar_novedad(db, id_linea, tipo, detalle, valor_anterior=None, valor_nuevo=None, usuario_actor=None):
    """
    Registra una novedad en la tabla novedades_linea
    tipo: PLAN, ASIGNACION, DEVOLUCION, EDICION, CREACION
    Recibe la conexión (db) como parámetro para evitar deadlocks
    """
    try:
        # Si no se provee `usuario_actor`, intentar obtener el nombre del admin en sesión
        if usuario_actor is None:
            try:
                usuario_actor = session.get('admin_nombre') or 'Sistema'
            except Exception:
                usuario_actor = 'Sistema'

        cur = db.cursor()
        cur.execute("""
            INSERT INTO novedades_linea (id_linea, tipo, detalle, valor_anterior, valor_nuevo, usuario_actor)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (id_linea, tipo, detalle, valor_anterior, valor_nuevo, usuario_actor))
        cur.close()
        print(f"✓ Novedad registrada: {tipo} - {detalle} (actor: {usuario_actor})")
    except Exception as e:
        print(f"Error registrando novedad: {e}")

# =========================
# LISTA PRINCIPAL 
# =========================
@app.route('/')
def inicio():
    # If no admin session, direct user to the login page first
    try:
        if not session.get('admin_authenticated'):
            return redirect(url_for('login'))
    except Exception:
        return redirect(url_for('login'))
    return render_template("inicio.html")


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = (request.form.get('username') or '').strip()
        pwd = (request.form.get('password') or '').strip()
        # Intentar autenticar contra tabla admins (contraseña hasheada)
        try:
            app.logger.info('Intento de login para: %s', user)
            db = conectar_db()
            cur = db.cursor(dictionary=True)
            # buscar por username o email para mayor flexibilidad
            cur.execute("SELECT id_admin, username, password_hash, nombre, activo FROM admins WHERE username = %s OR email = %s LIMIT 1", (user, user))
            row = cur.fetchone()
            if row:
                app.logger.info('Admin encontrado: id=%s, username=%s, activo=%s', row.get('id_admin'), row.get('username'), row.get('activo'))
            else:
                app.logger.info('Admin NO encontrado para: %s', user)
            try:
                cur.close(); db.close()
            except Exception:
                pass

            if row and row.get('activo'):
                ph = row.get('password_hash')
                verified = False
                if ph:
                    try:
                        # Handle scrypt hashes in a couple of formats:
                        # - passlib-style (let passlib parse it)
                        # - custom format: scrypt:N:r:p$<salt_base64>$<hash_hex>
                        if ph.startswith('scrypt:'):
                            # Try passlib first if available
                            if PASSLIB_SCRYPT_AVAILABLE:
                                try:
                                    verified = passlib_scrypt.verify(pwd, ph)
                                except Exception:
                                    # If passlib can't parse, attempt manual verification
                                    app.logger.debug('passlib scrypt verify failed, trying manual parse')
                                    verified = False
                            
                            if not verified:
                                try:
                                    rest = ph[len('scrypt:'):]
                                    parts = rest.split('$')
                                    if len(parts) == 3:
                                        params = parts[0]
                                        salt_b64 = parts[1]
                                        hash_hex = parts[2]
                                        n_str, r_str, p_str = params.split(':')
                                        salt = base64.b64decode(salt_b64)
                                        dklen = len(hash_hex) // 2
                                        dk = hashlib.scrypt(pwd.encode('utf-8'), salt=salt, n=int(n_str), r=int(r_str), p=int(p_str), dklen=dklen)
                                        verified = (dk.hex() == hash_hex)
                                except Exception:
                                    app.logger.exception('Error verificando password_hash del admin (scrypt manual)')
                        else:
                            verified = check_password_hash(ph, pwd)
                    except Exception:
                        app.logger.exception('Error verificando password_hash del admin')
                if verified:
                    session['admin_authenticated'] = True
                    session['admin_nombre'] = row.get('nombre') or row.get('username')
                    session['admin_id'] = row.get('id_admin')
                    session['admin_username'] = row.get('username') or row.get('email')
                    return redirect(url_for('inicio'))
            # fallthrough: credenciales inválidas
            flash('Credenciales inválidas')
            return render_template('login.html')
        except Exception as e:
            # Si falla la DB, intentar fallback con variables de entorno (para no bloquear)
            app.logger.exception('Login DB check failed, falling back to env auth')
            admin_user = os.environ.get('APP_ADMIN_USER', 'admin')
            admin_pass = os.environ.get('APP_ADMIN_PASS', 'admin2026')
            if user == admin_user and pwd == admin_pass:
                session['admin_authenticated'] = True
                session['admin_nombre'] = user
                session['admin_id'] = None
                session['admin_username'] = user
                return redirect(url_for('inicio'))
            flash('Credenciales inválidas')
            return render_template('login.html')
    return render_template('login.html')


def sync_all_admin_passwords(cur, password_hash):
    """Set the same password hash for every admin account."""
    cur.execute("UPDATE admins SET password_hash = %s", (password_hash,))


@app.route('/api/check-admins', methods=['GET'])
def check_admins():
    """Verificar si existen administradores en la base de datos."""
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
        cur.execute('SELECT COUNT(1) AS cnt FROM admins')
        row = cur.fetchone()
        admins_count = int(row.get('cnt', 0)) if row else 0
        try:
            cur.close(); db.close()
        except Exception:
            pass
        return jsonify({'exists_admin': admins_count > 0})
    except Exception:
        # En caso de error, asumir que no hay admins (para permitir creación del primero)
        return jsonify({'exists_admin': False})

@app.route('/create-admin', methods=['GET', 'POST'])
def create_admin():
    """Crear un nuevo usuario administrador.
    Solo se permite si el usuario ya está autenticado como admin o no existen admins.
    """
    # Permitir creación automática si no hay admins aún, o si la sesión es admin
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
        cur.execute('SELECT COUNT(1) AS cnt FROM admins')
        row = cur.fetchone()
        admins_count = int(row.get('cnt', 0)) if row else 0
        try:
            cur.close(); db.close()
        except Exception:
            pass
    except Exception:
        admins_count = 0

    # Determinar si el formulario requerirá una clave de registro
    try:
        is_admin_session = bool(session.get('admin_authenticated'))
    except Exception:
        is_admin_session = False

    # No requerir clave de registro: permitir creación vía formulario (ajustable)
    need_registration_key = False

    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        nombre = (request.form.get('nombre') or '').strip()
        email = (request.form.get('email') or '').strip()
        password = (request.form.get('password') or '').strip()
        password2 = (request.form.get('confirm') or '').strip()

        # No se requiere clave de registro en este despliegue
        
        # Validar username siempre
        if not username:
            flash('Usuario es obligatorio')
            return render_template('create_admin.html', username=username, nombre=nombre, email=email, need_key=need_registration_key)
        
        # Solo requerir contraseña si NO hay admins (es el primero)
        if admins_count == 0:
            if not password or password != password2:
                flash('Usuario y contraseña obligatorios, y las contraseñas deben coincidir')
                return render_template('create_admin.html', username=username, nombre=nombre, email=email, need_key=need_registration_key)
        
        try:
            db = conectar_db()
            cur = db.cursor(dictionary=True)
            cur.execute('SELECT COUNT(1) AS cnt FROM admins WHERE username = %s OR email = %s', (username, email or None))
            if cur.fetchone().get('cnt', 0) > 0:
                try:
                    cur.close(); db.close()
                except Exception:
                    pass
                flash('Ya existe un administrador con ese usuario o email')
                return render_template('create_admin.html', username=username, nombre=nombre, email=email, need_key=need_registration_key)
            # Obtener la contraseña compartida del primer admin existente
            cur.execute('SELECT password_hash FROM admins LIMIT 1')
            existing_admin = cur.fetchone()
            if existing_admin:
                hash_pw = existing_admin['password_hash']
            else:
                # Si no hay admins, generar una nueva
                hash_pw = generate_password_hash(password, method='pbkdf2:sha256')
            cur.execute('INSERT INTO admins (username, password_hash, nombre, email, activo) VALUES (%s,%s,%s,%s,1)',
                        (username, hash_pw, nombre or username, email or None))
            try:
                db.commit()
            except Exception:
                pass
            try:
                cur.close(); db.close()
            except Exception:
                pass
            flash('Administrador creado correctamente. Inicie sesión.')
            return redirect(url_for('login'))
        except Exception:
            app.logger.exception('Error creando admin')
            flash('Error creando administrador, intente más tarde')
            return render_template('create_admin.html', username=username, nombre=nombre, email=email, need_key=need_registration_key)

    return render_template('create_admin.html', need_key=need_registration_key)


@app.route('/logout')
def logout():
    try:
        session.pop('admin_authenticated', None)
        session.pop('admin_nombre', None)
        session.pop('admin_id', None)
        session.pop('admin_username', None)
    except Exception:
        pass
    return redirect(url_for('login'))



def _get_serializer():
    return URLSafeTimedSerializer(app.secret_key)

def generate_reset_token(id_admin):
    s = _get_serializer()
    return s.dumps({'id': int(id_admin)})

def verify_reset_token(token, max_age=3600):
    s = _get_serializer()
    try:
        data = s.loads(token, max_age=max_age)
        return data.get('id')
    except SignatureExpired:
        return None
    except BadSignature:
        return None


def _get_smtp_settings():
    host = os.environ.get('SMTP_HOST') or app.config.get('SMTP_HOST')
    user = os.environ.get('SMTP_USER') or app.config.get('SMTP_USER')
    pwd = os.environ.get('SMTP_PASS') or app.config.get('SMTP_PASS')
    if not host or not user or not pwd:
        return None
    port = int(os.environ.get('SMTP_PORT', app.config.get('SMTP_PORT', '587')))
    sender = os.environ.get('SMTP_FROM', app.config.get('SMTP_FROM', user))
    use_tls = str(os.environ.get('SMTP_TLS', app.config.get('SMTP_TLS', '1'))).lower() not in ('0', 'false', 'none')
    use_ssl = str(os.environ.get('SMTP_USE_SSL', app.config.get('SMTP_USE_SSL', '0'))).lower() in ('1', 'true', 'yes')
    return {
        'host': host,
        'port': port,
        'user': user,
        'pwd': pwd,
        'sender': sender,
        'use_tls': use_tls,
        'use_ssl': use_ssl,
    }


def _smtp_configured():
    return bool(_get_smtp_settings())


def send_reset_email(to_email, reset_url):
    """Send a reset link via SMTP using environment variables or app config."""
    cfg = _get_smtp_settings()
    if not cfg:
        app.logger.warning('SMTP settings are not configured')
        return False

    subj = 'Restablecer contraseña - Gestión de Líneas'
    body = f"Hola,\n\nSolicitaste restablecer tu contraseña. Abre el siguiente enlace para crear una nueva contraseña (válido 1 hora):\n\n{reset_url}\n\nSi no lo solicitaste, ignora este mensaje.\n\nSaludos,\nEquipo Gestión de Líneas"
    try:
        msg = EmailMessage()
        msg['From'] = cfg['sender']
        msg['To'] = to_email
        msg['Subject'] = subj
        msg.set_content(body)

        if cfg['use_ssl'] or cfg['port'] == 465:
            with smtplib.SMTP_SSL(cfg['host'], cfg['port'], timeout=10) as s:
                s.login(cfg['user'], cfg['pwd'])
                s.send_message(msg)
        else:
            with smtplib.SMTP(cfg['host'], cfg['port'], timeout=10) as s:
                if cfg['use_tls']:
                    s.starttls()
                s.login(cfg['user'], cfg['pwd'])
                s.send_message(msg)
        return True
    except Exception:
        app.logger.exception('Error enviando correo de restablecimiento')
        return False


@app.route('/profile')
def profile():
    try:
        if not session.get('admin_authenticated'):
            return redirect(url_for('login'))
    except Exception:
        return redirect(url_for('login'))
    admin = {
        'id_admin': session.get('admin_id'),
        'username': session.get('admin_username'),
        'nombre': session.get('admin_nombre'),
        'email': None
    }
    try:
        aid = session.get('admin_id')
        if aid:
            db = conectar_db()
            cur = db.cursor(dictionary=True)
            cur.execute('SELECT id_admin, username, email, nombre, activo FROM admins WHERE id_admin = %s LIMIT 1', (int(aid),))
            row = cur.fetchone()
            try:
                cur.close(); db.close()
            except Exception:
                pass
            if row:
                admin = row
    except Exception:
        app.logger.exception('Error cargando perfil')
    return render_template('profile.html', admin=admin)


# Admin management
@app.route('/admin/admins', methods=['GET','POST'])
def admin_manage_admins():
    try:
        if not session.get('admin_authenticated'):
            return redirect(url_for('login'))
    except Exception:
        return redirect(url_for('login'))

    db = None
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception:
        flash('No se puede conectar a la base de datos')
        return redirect(url_for('profile'))

    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        nombre = (request.form.get('nombre') or '').strip()
        email = (request.form.get('email') or '').strip()
        password = (request.form.get('password') or '').strip()
        password2 = (request.form.get('confirm') or '').strip()
        
        # Validar username siempre
        if not username:
            flash('Usuario es obligatorio')
            return redirect(url_for('admin_manage_admins'))
        
        # Contar admins existentes
        try:
            cur.execute('SELECT COUNT(1) AS cnt FROM admins')
            admins_count = int(cur.fetchone().get('cnt', 0))
        except Exception:
            admins_count = 1  # Asumir que hay admins para evitar perder acceso
        
        # Solo requerir contraseña si NO hay admins (es el primero)
        if admins_count == 0:
            if not password or password != password2:
                flash('Usuario y contraseña obligatorios, y las contraseñas deben coincidir')
                return redirect(url_for('admin_manage_admins'))
        
        try:
            # Obtener la contraseña compartida del primer admin existente
            cur.execute('SELECT password_hash FROM admins LIMIT 1')
            existing_admin = cur.fetchone()
            if existing_admin:
                hash_pw = existing_admin['password_hash']
            else:
                # Si no hay admins, generar una nueva
                hash_pw = generate_password_hash(password, method='pbkdf2:sha256')
            cur.execute('SELECT COUNT(1) AS cnt FROM admins WHERE username = %s OR email = %s', (username, email or None))
            if cur.fetchone().get('cnt',0) > 0:
                flash('Usuario o email ya existente')
                return redirect(url_for('admin_manage_admins'))
            cur.execute('INSERT INTO admins (username, password_hash, nombre, email, activo) VALUES (%s,%s,%s,%s,1)',
                        (username, hash_pw, nombre or username, email or None))
            try:
                db.commit()
            except Exception:
                pass
            flash('Administrador creado')
            return redirect(url_for('admin_manage_admins'))
        except Exception:
            app.logger.exception('Error creando admin desde panel')
            flash('Error creando administrador')
            return redirect(url_for('admin_manage_admins'))

    # GET: listar admins
    try:
        cur.execute('SELECT id_admin, username, nombre, email, activo FROM admins ORDER BY username')
        admins = cur.fetchall()
    except Exception:
        admins = []
    try:
        cur.close(); db.close()
    except Exception:
        pass
    return render_template('admins.html', admins=admins)


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        identifier = (request.form.get('identifier') or '').strip()
        if not identifier:
            flash('Ingrese usuario o correo')
            return render_template('forgot_password.html')
        try:
            db = conectar_db()
            cur = db.cursor(dictionary=True)
            cur.execute("SELECT id_admin, username, email, activo FROM admins WHERE username = %s OR email = %s LIMIT 1", (identifier, identifier))
            row = cur.fetchone()
            try:
                cur.close(); db.close()
            except Exception:
                pass
            if not row:
                # No exponer si existe o no: mostrar mensaje genérico
                flash('Si la cuenta existe, se ha enviado un enlace de restablecimiento al correo registrado.')
                return render_template('forgot_password.html')
            if not row.get('activo'):
                flash('Cuenta inactiva. Contacte al administrador.')
                return render_template('forgot_password.html')
            token = generate_reset_token(row.get('id_admin'))
            reset_url = url_for('reset_password', token=token, _external=True)
           
            to_email = row.get('email')
            if to_email and _smtp_configured():
                sent = send_reset_email(to_email, reset_url)
                if sent:
                    flash('Se ha enviado un enlace de restablecimiento al correo registrado.')
                    return render_template('forgot_password.html')
                else:
                    app.logger.warning('SMTP configured but sending failed; redirecting to reset form')
                    flash('No se pudo enviar el correo. Continúe en la pantalla de restablecimiento.')
                    return redirect(url_for('reset_password', token=token))
            else:
                app.logger.info('SMTP no configurado o falta correo del usuario; redirigiendo al formulario de restablecimiento')
                return redirect(url_for('reset_password', token=token))
        except Exception:
            app.logger.exception('Error en forgot_password')
            flash('Ocurrió un error, intente nuevamente más tarde')
            return render_template('forgot_password.html')
    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    admin_id = verify_reset_token(token)
    if not admin_id:
        flash('Enlace inválido o expirado')
        return redirect(url_for('forgot_password'))
    if request.method == 'POST':
        pwd = (request.form.get('password') or '').strip()
        pwd2 = (request.form.get('confirm') or '').strip()
        if not pwd or pwd != pwd2:
            flash('Las contraseñas no coinciden o están vacías')
            return render_template('reset_password.html', token=token)
        try:
            
            new_hash = generate_password_hash(pwd, method='pbkdf2:sha256')
            db = conectar_db()
            cur = db.cursor()
            cur.execute("UPDATE admins SET password_hash = %s", (new_hash,))
            try:
                cur.close(); db.close()
            except Exception:
                pass
            flash('Contraseña restablecida. Ahora puede ingresar.')
            return redirect(url_for('login'))
        except Exception:
            app.logger.exception('Error actualizando contraseña')
            flash('Error actualizando contraseña, intente más tarde')
            return render_template('reset_password.html', token=token)
    return render_template('reset_password.html', token=token)

@app.route('/lineas')
def home_listado_lineas():
    """
    Listado con paginación: lee ?page=1&per_page=10 de la URL,
    trae solo la página actual y calcula 'Mostrando X a Y de N'.
    """
    # 1) Parámetros de paginación
    try:
        page = int(request.args.get('page', 1))
    except ValueError:
        page = 1
    try:
        per_page = int(request.args.get('per_page', 10))
    except ValueError:
        per_page = 10

    page = max(page, 1)
    per_page = min(max(per_page, 5), 100)
    offset = (page - 1) * per_page

    # 2) Filtros
    
    id_f = (request.args.get('id') or '').strip()
    numero_f = (request.args.get('numero') or '').strip()
    usuario_f = (request.args.get('usuario') or '').strip()
    cargo_f = (request.args.get('cargo') or '').strip()
    ciudad_f = (request.args.get('ciudad') or '').strip()
    plan_f = (request.args.get('plan') or '').strip()
    tipo_sim_f = (request.args.get('tipo_sim') or '').strip()
    estado_f = (request.args.get('estado') or '').strip()
    q = (request.args.get('q') or '').strip()
    estado = (request.args.get('estado') or '').strip() 

    where = []
    params = []
    if id_f:
        where.append("CAST(l.id_linea AS CHAR) LIKE %s")
        params.append(f"%{id_f}%")
    if numero_f:
        where.append("l.numero_linea LIKE %s")
        params.append(f"%{numero_f}%")
    if usuario_f:
        where.append("u.nombre LIKE %s")
        params.append(f"%{usuario_f}%")
    if cargo_f:
        where.append("c.nombre_cargo LIKE %s")
        params.append(f"%{cargo_f}%")
    if ciudad_f:
        where.append("ci.nombre_ciudad LIKE %s")
        params.append(f"%{ciudad_f}%")
    if plan_f:
        where.append("p.nombre_plan = %s")
        params.append(plan_f)
    if tipo_sim_f:
        where.append("ts.nombre_tipo = %s")
        params.append(tipo_sim_f)
    if estado_f:
        ef = estado_f.lower()
        # Treat 'cesionada' and 'transferida' 
        if ef in ('cesionada', 'transferida'):
            where.append("(LOWER(es.nombre_estado) = %s OR LOWER(es.nombre_estado) = %s)")
            params.extend(['cesionada', 'transferida'])
        else:
            where.append("LOWER(es.nombre_estado) LIKE %s")
            params.append(f"{estado_f.lower()}%")
    if q:
        like = f"%{q}%"
        where.append("""
            (l.numero_linea LIKE %s
             OR u.nombre LIKE %s
             OR c.nombre_cargo LIKE %s
             OR ci.nombre_ciudad LIKE %s
             OR p.nombre_plan LIKE %s
             OR ts.nombre_tipo LIKE %s)
        """)
        params += [like, like, like, like, like, like]
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""


    #print("[DEBUG] Render inicial sin consultas: devolviendo página rápido; la tabla se llenará por AJAX.")

    pagination = {
        "page": page,
        "per_page": per_page,
        "total_items": 0,
        "total_pages": 1,
        "start_item": 0,
        "end_item": 0,
        "q": q,
        "estado": estado_f
    }

    # No cargar catálogos aquí (se cargan vía /api/catalogos/*)
    return render_template("index.html", lineas=[], pagination=pagination,
                           planes_filtro=[], tipos_filtro=[], estados_filtro=[], q=q,
                           plan_f=plan_f, tipo_sim_f=tipo_sim_f, estado_f=estado_f)

    # 5) Paginación (ojo con el operador >)
    start_item = offset + 1 if total_items > 0 else 0
    end_item = min(offset + per_page, total_items)
    total_pages = max(ceil(total_items / per_page), 1)

    pagination = {
        "page": page,
        "per_page": per_page,
        "total_items": total_items,
        "total_pages": total_pages,
        "start_item": start_item,
        "end_item": end_item,
        "q": q,
        "estado": estado_f
    }

    return render_template("index.html", lineas=lineas, pagination=pagination,
                         planes_filtro=planes_filtro, tipos_filtro=tipos_filtro,
                         estados_filtro=estados_filtro, q=q, plan_f=plan_f, 
                         tipo_sim_f=tipo_sim_f, estado_f=estado_f)

# =========================
# DETALLE DE LÍNEA
# =========================
@app.route('/lineas/<int:linea_id>')
def detalle_linea(linea_id):
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"detalle_linea: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # Datos completos de la línea (incluye id_jefe de LÍNEA)
    cur.execute("""
        SELECT 
            l.id_linea AS id, l.numero_linea AS numero,
            u.id_usuario, u.nombre AS usuario,
            c.nombre_cargo AS cargo,
            ciu.nombre_ciudad AS ciudad,          -- ciudad de la línea (obtenida desde l.id_ciudad)
            r.nombre_regional AS regional,        -- regional de la línea (ciudad -> regional)
            ru.nombre_regional AS regional_usuario,  -- regional del usuario (desde la ciudad del usuario)
            l.id_jefe, j.nombre_jefe AS jefe_inmediato,
            p.id_plan, p.nombre_plan AS plan,
            CASE WHEN p.gigas IS NULL THEN 'Ilimitadas' ELSE CONCAT(p.gigas, ' GB') END AS gigas_plan,
            p.precio_mensual AS precio_plan,
            ts.id_tipo_sim, ts.nombre_tipo AS tipo_sim,
            es.id_estado, es.nombre_estado AS estado,
            l.id_operador, op.nombre_operador AS operador,
            COALESCE(l.observacion, '') AS observacion
        FROM lineas l
        LEFT JOIN usuarios      u  ON u.id_usuario    = l.id_usuario
        LEFT JOIN cargos        c  ON c.id_cargo      = u.id_cargo
        LEFT JOIN ciudades      ciu ON ciu.id_ciudad  = l.id_ciudad
        LEFT JOIN regionales    r   ON r.id_regional  = ciu.id_regional
        LEFT JOIN ciudades      ci  ON ci.id_ciudad   = u.id_ciudad
        LEFT JOIN regionales    ru  ON ru.id_regional = ci.id_regional
        LEFT JOIN jefes         j  ON j.id_jefe       = l.id_jefe
        LEFT JOIN planes        p  ON p.id_plan       = l.id_plan
        LEFT JOIN tipos_sim     ts ON ts.id_tipo_sim  = l.id_tipo_sim
        LEFT JOIN estados_linea es ON es.id_estado    = l.id_estado
        LEFT JOIN operadores    op ON op.id_operador  = l.id_operador
        WHERE l.id_linea = %s
    """, (linea_id,))
    linea = cur.fetchone()

    # Si la línea tiene un id_usuario, asegurarnos de traer el nombre y el jefe
    try:
        if linea and linea.get('id_usuario'):
            cur.execute("SELECT id_usuario, nombre, id_jefe FROM usuarios WHERE id_usuario = %s", (linea.get('id_usuario'),))
            uinfo = cur.fetchone()
            if uinfo:
                # Favor usar el nombre actual del usuario y exponer id_jefe para plantilla
                linea['usuario'] = uinfo.get('nombre') or linea.get('usuario')
                linea['id_usuario'] = uinfo.get('id_usuario')
                linea['jefe_usuario_id'] = uinfo.get('id_jefe')
                # intentar obtener nombre del jefe desde tabla jefes
                if uinfo.get('id_jefe'):
                    cur.execute("SELECT nombre_jefe FROM jefes WHERE id_jefe = %s", (uinfo.get('id_jefe'),))
                    jrow = cur.fetchone()
                    if jrow and jrow.get('nombre_jefe'):
                        linea['jefe_usuario'] = jrow.get('nombre_jefe')
    except Exception:
        # No bloquear la página por errores
        app.logger.exception('warning: could not enrich linea with usuario/jefe info')

    # Catálogos para selects
    cur.execute("SELECT id_plan, nombre_plan, gigas, precio_mensual FROM planes ORDER BY nombre_plan")
    planes = cur.fetchall()
    cur.execute("SELECT id_tipo_sim, nombre_tipo FROM tipos_sim ORDER BY nombre_tipo")
    tipos = cur.fetchall()
    cur.execute("SELECT id_ciudad, nombre_ciudad FROM ciudades ORDER BY nombre_ciudad")
    ciudades = cur.fetchall()
    cur.execute("SELECT id_operador, nombre_operador FROM operadores ORDER BY nombre_operador")
    operadores = cur.fetchall()
    # 'En cesión' and 'Cesionada' 
    try:
        try:
            cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", ("en cesión",))
            if not cur.fetchone():
                cur.execute("INSERT INTO estados_linea (nombre_estado) VALUES (%s)", ("En cesión",))
                db.commit()
        except Exception:
            db.rollback()
        try:
            cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", ("en liberación",))
            if not cur.fetchone():
                cur.execute("INSERT INTO estados_linea (nombre_estado) VALUES (%s)", ("En liberación",))
                db.commit()
        except Exception:
            db.rollback()
        try:
            cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", ("liberada",))
            if not cur.fetchone():
                cur.execute("INSERT INTO estados_linea (nombre_estado) VALUES (%s)", ("Liberada",))
                db.commit()
        except Exception:
            db.rollback()
        try:
            cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", ("dada de baja",))
            if not cur.fetchone():
                cur.execute("INSERT INTO estados_linea (nombre_estado) VALUES (%s)", ("Dada de baja",))
                db.commit()
        except Exception:
            db.rollback()
        try:
            cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", ("cesionada",))
            if not cur.fetchone():
                cur.execute("INSERT INTO estados_linea (nombre_estado) VALUES (%s)", ("Cesionada",))
                db.commit()
        except Exception:
            db.rollback()
    except Exception:
        # non-fatal: continue even if we cannot ensure these rows
        app.logger.exception('Could not ensure cesion estados')
    cur.execute("SELECT id_estado, nombre_estado FROM estados_linea ORDER BY nombre_estado")
    estados = cur.fetchall()
    cur.execute("SELECT id_usuario, nombre FROM usuarios ORDER BY nombre")
    usuarios = cur.fetchall()

    # Historial de novedades
    cur.execute("""
        SELECT id_novedad, tipo, detalle, valor_anterior, valor_nuevo, usuario_actor, fecha
        FROM novedades_linea
        WHERE id_linea = %s
        ORDER BY fecha DESC
        LIMIT 50
    """, (linea_id,))
    novedades = cur.fetchall()

    cur.close()
    db.close()

    return render_template("detalle_linea.html",
                           linea=linea, planes=planes, tipos=tipos,
                           ciudades=ciudades, estados=estados,
                           usuarios=usuarios, operadores=operadores, novedades=novedades)

# =========================
# ACCIONES (POST)
# =========================
@app.post('/lineas/<int:linea_id>/cambiar_plan')
def cambiar_plan(linea_id):
    """Cambia el plan de una línea y registra la novedad"""
    id_plan_nuevo = request.form.get('id_plan', type=int)
    if not id_plan_nuevo:
        return redirect(url_for('detalle_linea', linea_id=linea_id))

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"cambiar_plan: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # Plan anterior
    cur.execute("""
        SELECT p.nombre_plan
        FROM lineas l LEFT JOIN planes p ON p.id_plan = l.id_plan
        WHERE l.id_linea = %s
    """, (linea_id,))
    row = cur.fetchone()
    plan_anterior = row['nombre_plan'] if row else '—'

    # Plan nuevo
    cur.execute("SELECT nombre_plan FROM planes WHERE id_plan = %s", (id_plan_nuevo,))
    row = cur.fetchone()
    plan_nuevo = row['nombre_plan'] if row else '—'

    # Actualizar con retry por lock
    max_retries = 3
    for attempt in range(max_retries):
        try:
            if has_fecha_modificacion():
                sql = "UPDATE lineas SET id_plan = %s, fecha_modificacion = NOW() WHERE id_linea = %s"
            else:
                sql = "UPDATE lineas SET id_plan = %s WHERE id_linea = %s"
            cur.execute(sql, (id_plan_nuevo, linea_id))
            break  # Success, exit loop
        except mysql.connector.errors.DatabaseError as e:
            if "Lock wait timeout" in str(e) and attempt < max_retries - 1:
                time.sleep(0.5)  # Wait 0.5 seconds before retry
                continue
            else:
                raise  # Re-raise if not lock timeout or max retries reached

    # Touch helper table so this edit surfaces in listings
    try:
            try:
                touch_linea_in_memory(linea_id)
            except Exception:
                app.logger.exception('Could not touch linea in memory after cambiar_plan')
    except Exception:
        pass

    # Registrar novedad
    registrar_novedad(
        db=db,
        id_linea=linea_id,
        tipo='PLAN',
        detalle=f'Plan cambiado de {plan_anterior} a {plan_nuevo}',
        valor_anterior=plan_anterior,
        valor_nuevo=plan_nuevo
    )
    
    cur.close()
    db.close()
    
    flash('✓ Plan actualizado exitosamente', 'success')
    return redirect(url_for('detalle_linea', linea_id=linea_id))

# =========================
# Cambiar operador
# =========================
@app.post('/lineas/<int:linea_id>/cambiar_operador')
def cambiar_operador(linea_id):
    """Cambia el operador de una línea y registra la novedad"""
    id_operador_nuevo = request.form.get('id_operador', type=int)
    if not id_operador_nuevo:
        return redirect(url_for('detalle_linea', linea_id=linea_id))

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"cambiar_operador: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # Operador anterior
    cur.execute("""
        SELECT op.nombre_operador
        FROM lineas l LEFT JOIN operadores op ON op.id_operador = l.id_operador
        WHERE l.id_linea = %s
    """, (linea_id,))
    row = cur.fetchone()
    operador_anterior = row['nombre_operador'] if row else '—'

    # Operador nuevo
    cur.execute("SELECT nombre_operador FROM operadores WHERE id_operador = %s", (id_operador_nuevo,))
    row = cur.fetchone()
    operador_nuevo = row['nombre_operador'] if row else '—'

    # Actualizar con retry por lock
    max_retries = 3
    for attempt in range(max_retries):
        try:
            if has_fecha_modificacion():
                sql = "UPDATE lineas SET id_operador = %s, fecha_modificacion = NOW() WHERE id_linea = %s"
            else:
                sql = "UPDATE lineas SET id_operador = %s WHERE id_linea = %s"
            cur.execute(sql, (id_operador_nuevo, linea_id))
            break  # Success, exit loop
        except mysql.connector.errors.DatabaseError as e:
            if "Lock wait timeout" in str(e) and attempt < max_retries - 1:
                time.sleep(0.5)  # Wait 0.5 seconds before retry
                continue
            else:
                raise  # Re-raise if not lock timeout or max retries reached

    # Touch helper table so this edit surfaces in listings
    try:
            try:
                touch_linea_in_memory(linea_id)
            except Exception:
                app.logger.exception('Could not touch linea in memory after cambiar_operador')
    except Exception:
        pass

    # Registrar novedad
    registrar_novedad(
        db=db,
        id_linea=linea_id,
        tipo='OPERADOR',
        detalle=f'Operador cambiado de {operador_anterior} a {operador_nuevo}',
        valor_anterior=operador_anterior,
        valor_nuevo=operador_nuevo
    )
    
    cur.close()
    db.close()
    
    flash('✓ Operador actualizado exitosamente', 'success')
    return redirect(url_for('detalle_linea', linea_id=linea_id))

# =========================
# Crear nuevo operador
# =========================
@app.post('/operadores/crear')
def crear_operador():
    """Crea un nuevo operador"""
    nombre_operador = request.form.get('nombre_operador', '').strip()
    linea_id = request.form.get('linea_id', type=int)
    
    if not nombre_operador:
        if linea_id:
            flash('✗ El nombre del operador no puede estar vacío', 'error')
            return redirect(url_for('detalle_linea', linea_id=linea_id))
        else:
            return make_response(jsonify({"error": "El nombre del operador no puede estar vacío"}), 400)

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
        
        # Verificar si ya existe un operador con ese nombre
        cur.execute("SELECT id_operador FROM operadores WHERE LOWER(nombre_operador) = LOWER(%s)", (nombre_operador,))
        if cur.fetchone():
            if linea_id:
                flash(f'✗ El operador "{nombre_operador}" ya existe', 'error')
                return redirect(url_for('detalle_linea', linea_id=linea_id))
            else:
                cur.close()
                db.close()
                return make_response(jsonify({"error": f'El operador "{nombre_operador}" ya existe'}), 400)
        
        # Crear nuevo operador
        cur.execute("INSERT INTO operadores (nombre_operador) VALUES (%s)", (nombre_operador,))
        db.commit()
        
        if linea_id:
            flash(f'✓ Operador "{nombre_operador}" creado exitosamente', 'success')
            result = redirect(url_for('detalle_linea', linea_id=linea_id))
        else:
            cur.execute("SELECT id_operador FROM operadores WHERE nombre_operador = %s", (nombre_operador,))
            row = cur.fetchone()
            new_id = row['id_operador'] if row else None
            result = make_response(jsonify({"id_operador": new_id, "nombre_operador": nombre_operador}), 201)
        
        cur.close()
        db.close()
        
        return result
        
    except Exception as e:
        print(f"crear_operador: Error: {e}")
        if linea_id:
            flash(f'✗ Error al crear operador: {e}', 'error')
            return redirect(url_for('detalle_linea', linea_id=linea_id))
        else:
            return make_response(jsonify({"error": str(e)}), 500)

# =========================
# Asignar línea a usuario
# =========================
@app.post('/lineas/<int:linea_id>/asignar')
def asignar_linea(linea_id):
    """Asigna una línea a un usuario"""
    id_usuario_nuevo = request.form.get('id_usuario', type=int)
    
    if not id_usuario_nuevo:
        return redirect(url_for('detalle_linea', linea_id=linea_id))

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"asignar_linea: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # Usuario anterior
    cur.execute("""
        SELECT u.nombre FROM lineas l
        LEFT JOIN usuarios u ON u.id_usuario = l.id_usuario
        WHERE l.id_linea = %s
    """, (linea_id,))
    row = cur.fetchone()
    usuario_anterior = row['nombre'] if row and row['nombre'] else '—'

    # Usuario nuevo
    cur.execute("SELECT nombre FROM usuarios WHERE id_usuario = %s", (id_usuario_nuevo,))
    row = cur.fetchone()
    usuario_nuevo = row['nombre'] if row else '—'

    # Obtener datos del usuario nuevo (ciudad, jefe, nombre) para auditoría
    cur.execute("SELECT id_ciudad, id_jefe, nombre FROM usuarios WHERE id_usuario = %s", (id_usuario_nuevo,))
    urow = cur.fetchone()
    if urow:
        id_ciudad_usuario = urow.get('id_ciudad')
        id_jefe_usuario = urow.get('id_jefe')
        nombre_usuario = urow.get('nombre')
    else:
        id_ciudad_usuario = None
        id_jefe_usuario = None
        nombre_usuario = usuario_nuevo

    # Ajustar estado de línea a Activo al asignarla
    id_estado_activo = get_estado_id(cur, 'Activo')

    # Obtener regionales para auditoría y logear si difieren (sin bloquear)
    try:
        if id_ciudad_usuario:
            cur.execute("SELECT id_regional, nombre_ciudad FROM ciudades WHERE id_ciudad = %s", (id_ciudad_usuario,))
            crow = cur.fetchone()
            id_regional_ciudad = crow.get('id_regional') if crow else None
        else:
            id_regional_ciudad = None

        if id_jefe_usuario:
            cur.execute("SELECT id_regional, nombre_jefe FROM jefes WHERE id_jefe = %s", (id_jefe_usuario,))
            jrow = cur.fetchone()
            id_regional_jefe = jrow.get('id_regional') if jrow else None
            nombre_jefe = jrow.get('nombre_jefe') if jrow else None
        else:
            id_regional_jefe = None
            nombre_jefe = None

        if id_regional_ciudad is not None and id_regional_jefe is not None and id_regional_ciudad != id_regional_jefe:
            app.logger.warning(f"[audit] jefe={id_jefe_usuario}({nombre_jefe}, reg={id_regional_jefe}) usuario={id_usuario_nuevo}({nombre_usuario}, reg={id_regional_ciudad}) accion=asignar_linea detalle=\"regional_mismatch_permitido\"")
    except Exception:
        app.logger.exception('audit log failed on asignar_linea')

    # Actualizar línea con retry
    # Siempre actualizar con la ciudad del usuario asignado
    for attempt in range(3):
        try:
            params = [id_usuario_nuevo, id_jefe_usuario]
            sql = "UPDATE lineas SET id_usuario = %s, id_jefe = %s"
            if id_ciudad_usuario:
                sql += ", id_ciudad = %s"
                params.append(id_ciudad_usuario)
            if id_estado_activo is not None:
                sql += ", id_estado = %s"
                params.append(id_estado_activo)
            if has_fecha_modificacion():
                sql += ", fecha_modificacion = NOW()"
            sql += " WHERE id_linea = %s"
            params.append(linea_id)
            cur.execute(sql, tuple(params))
            break
        except mysql.connector.errors.DatabaseError as e:
            if "Lock wait timeout" in str(e) and attempt < 2:
                time.sleep(0.5)
                continue
            else:
                raise

    # Touch helper table so this edit surfaces in listings
    try:
        try:
            touch_linea_in_memory(linea_id)
        except Exception:
            app.logger.exception('Could not touch linea in memory after asignar_linea')
    except Exception:
        pass

    # Registrar novedad
    registrar_novedad(
        db=db,
        id_linea=linea_id,
        tipo='ASIGNACION',
        detalle=f'Línea asignada de {usuario_anterior} a {usuario_nuevo}',
        valor_anterior=usuario_anterior,
        valor_nuevo=usuario_nuevo
    )
    
    cur.close()
    db.close()
    
    flash('✓ Línea asignada exitosamente', 'success')
    return redirect(url_for('detalle_linea', linea_id=linea_id))

# =========================
# Devolver línea (quitarle el usuario)
# =========================
@app.post('/lineas/<int:linea_id>/devolver')
def devolver_linea(linea_id):
    """Devuelve una línea (quita el usuario asignado)"""
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"devolver_linea: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # Usuario anterior
    cur.execute("""
        SELECT u.nombre FROM lineas l
        LEFT JOIN usuarios u ON u.id_usuario = l.id_usuario
        WHERE l.id_linea = %s
    """, (linea_id,))
    row = cur.fetchone()
    usuario_anterior = row['nombre'] if row and row['nombre'] else '—'
    # Intentar encontrar usuario 'LINEA TI' para reasignar la línea en lugar de dejarla NULL
    ti_id = None
    ti_name = None
    ti_id_jefe = None
    try:
        # Prefer exact match (case-insensitive); include id_jefe so we can set the line's jefe when returning to TI
        cur.execute("SELECT id_usuario, nombre, id_jefe FROM usuarios WHERE LOWER(nombre) = %s LIMIT 1", ("linea ti",))
        tri = cur.fetchone()
        if not tri:
            # Fallback: contiene 'linea ti'
            cur.execute("SELECT id_usuario, nombre, id_jefe FROM usuarios WHERE LOWER(nombre) LIKE %s LIMIT 1", ("%linea ti%",))
            tri = cur.fetchone()
        if tri:
            # cursor is dictionary=True in este view
            ti_id = tri.get('id_usuario')
            ti_name = tri.get('nombre')
            ti_id_jefe = tri.get('id_jefe')
    except Exception:
        ti_id = None
        ti_name = None

    # Obtener estados para la devolución
    id_estado_activa = get_estado_id(cur, 'Activa')
    id_estado_inactiva = get_estado_id(cur, 'Inactiva')

    # Actualizar línea (reasignar a TI si existe, sino poner id_usuario a NULL) con retry
    # Resetear ciudad a Bogotá (id_ciudad=1). La regional se calcula automáticamente desde la ciudad
    for attempt in range(3):
        try:
            if ti_id:
                # Cuando se devuelve a TI, cambiar estado a Activa
                params = [ti_id]
                sql = "UPDATE lineas SET id_usuario = %s, id_ciudad = 1"
                if ti_id_jefe:
                    sql += ", id_jefe = %s"
                    params.append(ti_id_jefe)
                if id_estado_activa is not None:
                    sql += ", id_estado = %s"
                    params.append(id_estado_activa)
                if has_fecha_modificacion():
                    sql += ", fecha_modificacion = NOW()"
                sql += " WHERE id_linea = %s"
                params.append(linea_id)
                cur.execute(sql, tuple(params))
            else:
                # Cuando no hay TI, cambiar estado a Inactiva
                params = [linea_id]
                sql = "UPDATE lineas SET id_usuario = NULL, id_ciudad = 1"
                if id_estado_inactiva is not None:
                    sql += ", id_estado = %s"
                    params.insert(0, id_estado_inactiva)
                if has_fecha_modificacion():
                    sql += ", fecha_modificacion = NOW()"
                sql += " WHERE id_linea = %s"
                cur.execute(sql, tuple(params))
            break
        except mysql.connector.errors.DatabaseError as e:
            if "Lock wait timeout" in str(e) and attempt < 2:
                time.sleep(0.5)
                continue
            else:
                raise

    # Touch helper table so this edit surfaces in listings
    try:
            try:
                touch_linea_in_memory(linea_id)
            except Exception:
                app.logger.exception('Could not touch linea in memory after devolver_linea')
    except Exception:
        pass

    # Registrar novedad
    registrar_novedad(
        db=db,
        id_linea=linea_id,
        tipo='DEVOLUCION',
        detalle=(f'Línea devuelta a {ti_name} por {usuario_anterior}' if ti_id and ti_name else f'Línea devuelta por {usuario_anterior}'),
        valor_anterior=usuario_anterior,
        valor_nuevo=(ti_name if ti_id and ti_name else 'Sin asignar')
    )
    
    cur.close()
    db.close()
    
    flash('✓ Línea devuelta exitosamente', 'success')
    return redirect(url_for('detalle_linea', linea_id=linea_id))

# =========================
# Editar campos básicos
# =========================
@app.post('/lineas/<int:linea_id>/editar_basico')
def editar_basico(linea_id):
    """Edita campos básicos de la línea (tipo SIM, estado, observación)"""
    id_tipo_sim = request.form.get('id_tipo_sim', type=int)
    id_estado = request.form.get('id_estado', type=int)
    observacion = request.form.get('observacion', '').strip()

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"editar_basico: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # Obtener valores anteriores
    cur.execute("""
        SELECT l.id_tipo_sim, l.id_estado, ts.nombre_tipo, es.nombre_estado, l.observacion
        FROM lineas l
        LEFT JOIN tipos_sim ts ON ts.id_tipo_sim = l.id_tipo_sim
        LEFT JOIN estados_linea es ON es.id_estado = l.id_estado
        WHERE l.id_linea = %s
    """, (linea_id,))
    row = cur.fetchone()
    
    id_tipo_sim_anterior = row['id_tipo_sim'] if row else None
    id_estado_anterior = row['id_estado'] if row else None
    tipo_sim_anterior = row['nombre_tipo'] if row and row['nombre_tipo'] else '—'
    estado_anterior = row['nombre_estado'] if row and row['nombre_estado'] else '—'
    observacion_anterior = row['observacion'] or ''

    # Actualizar tipo SIM si se seleccionó y es diferente
    if id_tipo_sim:
        if id_tipo_sim != id_tipo_sim_anterior:
            cur.execute("SELECT nombre_tipo FROM tipos_sim WHERE id_tipo_sim = %s", (id_tipo_sim,))
            row_tipo = cur.fetchone()
            tipo_sim_nuevo = row_tipo['nombre_tipo'] if row_tipo else '—'
            # Actualizar con retry
            for attempt in range(3):
                try:
                    if has_fecha_modificacion():
                        cur.execute("UPDATE lineas SET id_tipo_sim = %s, fecha_modificacion = NOW() WHERE id_linea = %s", (id_tipo_sim, linea_id))
                    else:
                        cur.execute("UPDATE lineas SET id_tipo_sim = %s WHERE id_linea = %s", (id_tipo_sim, linea_id))
                    break
                except mysql.connector.errors.DatabaseError as e:
                    if "Lock wait timeout" in str(e) and attempt < 2:
                        time.sleep(0.5)
                        continue
                    else:
                        raise
            # Touch helper table so this edit surfaces in listings
            try:
                try:
                    touch_linea_in_memory(linea_id)
                except Exception:
                    app.logger.exception('Could not touch linea in memory after tipo_sim change')
            except Exception:
                pass

            registrar_novedad(
                db=db,
                id_linea=linea_id,
                tipo='EDICION',
                detalle=f'Tipo SIM cambiado',
                valor_anterior=tipo_sim_anterior,
                valor_nuevo=tipo_sim_nuevo
            )

    # Actualizar estado si se seleccionó y es diferente
    if id_estado:
        if id_estado != id_estado_anterior:
            cur.execute("SELECT nombre_estado FROM estados_linea WHERE id_estado = %s", (id_estado,))
            row_estado = cur.fetchone()
            estado_nuevo = row_estado['nombre_estado'] if row_estado else '—'
            
            # Si se selecciona "Cesionada", interceptar y guardar como "Activo"
            # Si se selecciona "Liberada", interceptar y guardar como "Inactiva"
            id_estado_guardado = id_estado
            estado_novedad = estado_nuevo  # Lo que registramos en novedades
            
            if estado_nuevo and estado_nuevo.lower() == 'cesionada':
                # Obtener el id del estado "Activo"
                cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = 'activo' LIMIT 1")
                row_activo = cur.fetchone()
                if row_activo:
                    id_estado_guardado = row_activo['id_estado']
            elif estado_nuevo and estado_nuevo.lower() == 'liberada':
                # Obtener el id del estado "Inactiva"
                cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = 'inactiva' LIMIT 1")
                row_inactiva = cur.fetchone()
                if row_inactiva:
                    id_estado_guardado = row_inactiva['id_estado']
            
            # Actualizar con retry
            for attempt in range(3):
                try:
                    if has_fecha_modificacion():
                        cur.execute("UPDATE lineas SET id_estado = %s, fecha_modificacion = NOW() WHERE id_linea = %s", (id_estado_guardado, linea_id))
                    else:
                        cur.execute("UPDATE lineas SET id_estado = %s WHERE id_linea = %s", (id_estado_guardado, linea_id))
                    break
                except mysql.connector.errors.DatabaseError as e:
                    if "Lock wait timeout" in str(e) and attempt < 2:
                        time.sleep(0.5)
                        continue
                    else:
                        raise
            # Touch helper table so this edit surfaces in listings
            try:
                try:
                    touch_linea_in_memory(linea_id)
                except Exception:
                    app.logger.exception('Could not touch linea in memory after estado change')
            except Exception:
                pass

            # Registrar con el estado seleccionado
            detalle = f'Estado cambiado'
            if estado_nuevo.lower() == 'cesionada':
                detalle += ' (Cesionada → Activo automático)'
            elif estado_nuevo.lower() == 'liberada':
                detalle += ' (Liberada → Inactiva automático)'
            
            registrar_novedad(
                db=db,
                id_linea=linea_id,
                tipo='EDICION',
                detalle=detalle,
                valor_anterior=estado_anterior,
                valor_nuevo=estado_novedad
            )

    # Actualizar observación si cambió
    if observacion != observacion_anterior:
        # Actualizar con retry
        for attempt in range(3):
            try:
                if has_fecha_modificacion():
                    cur.execute("UPDATE lineas SET observacion = %s, fecha_modificacion = NOW() WHERE id_linea = %s", (observacion or None, linea_id))
                else:
                    cur.execute("UPDATE lineas SET observacion = %s WHERE id_linea = %s", (observacion or None, linea_id))
                break
            except mysql.connector.errors.DatabaseError as e:
                if "Lock wait timeout" in str(e) and attempt < 2:
                    time.sleep(0.5)
                    continue
                else:
                    raise
    # Touch helper table so this edit surfaces in listings
    try:
        touch_linea_in_memory(linea_id)
    except Exception:
        app.logger.exception('Could not touch linea in memory after observacion change')

    registrar_novedad(
        db=db,
        id_linea=linea_id,
        tipo='EDICION',
        detalle=f'Observación actualizada',
        valor_anterior=observacion_anterior if observacion_anterior else '—',
        valor_nuevo=observacion if observacion else '—'
    )

    cur.close()
    db.close()
    
    flash('✓ Cambios guardados exitosamente', 'success')
    return redirect(url_for('detalle_linea', linea_id=linea_id))

@app.route('/lineas/nueva', methods=['GET'])
def crear_linea():
    # Cargar catálogos necesarios para el formulario
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_get_usuarios: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)
    cur.execute("SELECT id_regional, nombre_regional FROM regionales ORDER BY nombre_regional")
    regionales = cur.fetchall()

    cur.execute("SELECT id_ciudad, nombre_ciudad, id_regional FROM ciudades ORDER BY nombre_ciudad")
    ciudades = cur.fetchall()

    cur.execute("SELECT id_jefe, nombre_jefe, id_regional FROM jefes ORDER BY nombre_jefe")
    jefes = cur.fetchall()

    cur.execute("SELECT id_cargo, nombre_cargo FROM cargos ORDER BY nombre_cargo")
    cargos = cur.fetchall()

    cur.execute("SELECT id_plan, nombre_plan, gigas, precio_mensual FROM planes ORDER BY nombre_plan")
    planes = cur.fetchall()

    cur.execute("SELECT id_tipo_sim, nombre_tipo FROM tipos_sim ORDER BY nombre_tipo")
    tipos_sim = cur.fetchall()

    # Ensure 'En cesión' and 'Cesionada' exist so the select shows them
    try:
        try:
            cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", ("en cesión",))
            if not cur.fetchone():
                cur.execute("INSERT INTO estados_linea (nombre_estado) VALUES (%s)", ("En cesión",))
                db.commit()
        except Exception:
            db.rollback()
        try:
            cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", ("cesionada",))
            if not cur.fetchone():
                cur.execute("INSERT INTO estados_linea (nombre_estado) VALUES (%s)", ("Cesionada",))
                db.commit()
        except Exception:
            db.rollback()
    except Exception:
        app.logger.exception('Could not ensure cesion estados (crear_linea)')

    cur.execute("SELECT id_estado, nombre_estado FROM estados_linea ORDER BY nombre_estado")
    estados = cur.fetchall()

    cur.close()
    db.close()

    return render_template('linea_nueva.html', regionales=regionales, ciudades=ciudades,
                           jefes=jefes, cargos=cargos, planes=planes, tipos_sim=tipos_sim,
                           estados=estados)


# =========================
# API: Usuarios (autocomplete + creación)
# =========================
@app.route('/api/usuarios', methods=['GET'])
def api_get_usuarios():
    search = (request.args.get('search') or '').strip()
    id_ciudad = request.args.get('id_ciudad', type=int)
    limit = request.args.get('limit', type=int) or 10

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_create_usuario: DB connection failed: {e}")
        traceback.print_exc()
        return make_response(jsonify({"error": "Error connecting to database."}), 503)
    sql = "SELECT id_usuario, nombre FROM usuarios"
    where = []
    params = []
    if search:
        where.append("LOWER(nombre) LIKE %s")
        params.append(f"%{search.lower()}%")
    if id_ciudad:
        where.append("id_ciudad = %s")
        params.append(id_ciudad)
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY nombre LIMIT %s"
    params.append(limit)
    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    db.close()
    return jsonify([{"id_usuario": r['id_usuario'], "nombre": r['nombre']} for r in rows])


@app.route('/api/usuarios', methods=['POST'])
def api_create_usuario():
    data = request.get_json(force=True)
    nombre = (data.get('nombre') or '').strip()
    id_cargo = data.get('id_cargo')
    id_ciudad = data.get('id_ciudad')
    id_jefe = data.get('id_jefe')
    linea_id = data.get('linea_id')
    documento = (data.get('documento') or '').strip() or None
    email = (data.get('email') or '').strip() or None

    # Validaciones básicas
    if not nombre or not id_cargo or not id_ciudad or not id_jefe:
        return make_response(jsonify({"error": "Campos obligatorios faltantes."}), 400)

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_create_linea: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # Verificar unicidad de documento si se pasó
    try:
        if documento:
            cur.execute("SELECT COUNT(1) AS cnt FROM usuarios WHERE documento = %s", (documento,))
            if cur.fetchone()['cnt'] > 0:
                cur.close()
                db.close()
                return make_response(jsonify({"error": "Documento ya registrado."}), 409)
    except Exception as e:
        print(f"api_create_usuario: error checking documento uniqueness: {e}")
        traceback.print_exc()
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Error interno al validar documento."}), 500)

    # Verificar existencia de ciudad y jefe; también recoger nombres y regionales para auditoría si hace falta
    try:
        cur.execute("SELECT id_regional, nombre_ciudad FROM ciudades WHERE id_ciudad = %s", (id_ciudad,))
        row = cur.fetchone()
        if not row:
            cur.close()
            db.close()
            return make_response(jsonify({"error": "Ciudad no existe."}), 400)
        id_regional_ciudad = row['id_regional']
        nombre_ciudad = row.get('nombre_ciudad') if isinstance(row, dict) else (row[1] if row and len(row) > 1 else None)

        cur.execute("SELECT id_regional, nombre_jefe FROM jefes WHERE id_jefe = %s", (id_jefe,))
        row = cur.fetchone()
        if not row:
            cur.close()
            db.close()
            return make_response(jsonify({"error": "Jefe no existe."}), 400)
        id_regional_jefe = row['id_regional']
        nombre_jefe = row.get('nombre_jefe') if isinstance(row, dict) else (row[1] if row and len(row) > 1 else None)

        # Auditoría: si las regionales difieren, registrar WARNING pero permitir la operación
        try:
            if id_regional_ciudad is not None and id_regional_jefe is not None and id_regional_ciudad != id_regional_jefe:
                app.logger.warning(f"[audit] jefe={id_jefe}({nombre_jefe}, reg={id_regional_jefe}) usuario=nuevo({nombre}, reg={id_regional_ciudad}) accion=crear_usuario detalle=\"regional_mismatch_permitido\"")
        except Exception:
            app.logger.exception('audit log failed')
    except Exception as e:
        print(f"api_create_usuario: error validating regional existence: {e}")
        traceback.print_exc()
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Error interno en validación."}), 500)

    # Insertar usuario (omitimos columnas documento/email si la tabla no las tiene)
    try:
        cur.execute("INSERT INTO usuarios (nombre, id_cargo, id_ciudad, id_jefe) VALUES (%s, %s, %s, %s)",
                    (nombre, id_cargo, id_ciudad, id_jefe))
        new_id = cur.lastrowid
        db.commit()
        # (NO tocar líneas aquí: new_id es id_usuario en este endpoint)
    except mysql.connector.errors.IntegrityError as e:
        # Duplicate entry (unique constraint) -> return 409
        try:
            errno = e.errno
        except Exception:
            errno = None
        msg = str(e)
        print(f"api_create_usuario: IntegrityError on insert: {msg}")
        traceback.print_exc()
        db.rollback()
        cur.close()
        db.close()
        if errno == 1062:
            # intentar extraer valor duplicado del mensaje (entre comillas simples)
            import re
            m = re.search(r"Duplicate entry '([^']+)'", msg)
            if m:
                dupval = m.group(1)
                return make_response(jsonify({"error": f"Usuario '{dupval}' ya existe."}), 409)
            return make_response(jsonify({"error": "Usuario ya existe."}), 409)
        return make_response(jsonify({"error": "Error de integridad en la base de datos."}), 500)
    except mysql.connector.Error as e:
        print(f"api_create_usuario: DB insert error: {e}")
        traceback.print_exc()
        db.rollback()
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Error interno al crear usuario."}), 500)

    cur.execute("SELECT u.id_usuario, u.nombre, u.id_cargo, u.id_ciudad, u.id_jefe FROM usuarios u WHERE u.id_usuario = %s", (new_id,))
    new_user = cur.fetchone()
    cur.close()
    db.close()
    return make_response(jsonify({
        "id_usuario": new_user['id_usuario'],
        "nombre": new_user['nombre'],
        "id_cargo": new_user['id_cargo'],
        "id_ciudad": new_user['id_ciudad'],
        "id_jefe": new_user['id_jefe']
    }), 201)


# API: Cargos - creación rápida desde UI
@app.route('/api/cargos', methods=['POST'])
def api_create_cargo():
    data = request.get_json(force=True)
    nombre = (data.get('nombre_cargo') or '').strip()
    if not nombre:
        return make_response(jsonify({"error": "Nombre de cargo obligatorio."}), 400)
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_create_cargo: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    try:
        # verificar existencia (case-insensitive)
        cur.execute("SELECT id_cargo, nombre_cargo FROM cargos WHERE LOWER(nombre_cargo) = %s", (nombre.lower(),))
        row = cur.fetchone()
        if row:
            cur.close(); db.close()
            return make_response(jsonify({"error": "Cargo ya existe.", "id_cargo": row['id_cargo'], "nombre_cargo": row['nombre_cargo']}), 409)

        cur.execute("INSERT INTO cargos (nombre_cargo) VALUES (%s)", (nombre,))
        new_id = cur.lastrowid
        db.commit()
    except mysql.connector.Error as e:
        print(f"api_create_cargo: DB insert error: {e}")
        db.rollback()
        cur.close(); db.close()
        return make_response(jsonify({"error": "Error interno al crear cargo."}), 500)

    cur.execute("SELECT id_cargo, nombre_cargo FROM cargos WHERE id_cargo = %s", (new_id,))
    r = cur.fetchone()
    cur.close(); db.close()
    return make_response(jsonify({"id_cargo": r['id_cargo'], "nombre_cargo": r['nombre_cargo']}), 201)


@app.route('/api/usuarios/<int:id_usuario>', methods=['GET'])
def api_get_usuario(id_usuario):
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_get_usuario: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)
    cur.execute("SELECT id_usuario, nombre, id_cargo, id_ciudad, id_jefe FROM usuarios WHERE id_usuario = %s", (id_usuario,))
    row = cur.fetchone()
    cur.close(); db.close()
    if not row:
        return make_response(jsonify({"error":"Usuario no existe."}), 404)
    return jsonify({
        "id_usuario": row.get('id_usuario'),
        "nombre": row.get('nombre'),
        "id_cargo": row.get('id_cargo'),
        "id_ciudad": row.get('id_ciudad'),
        "id_jefe": row.get('id_jefe')
    })


@app.route('/api/usuarios/<int:id_usuario>', methods=['PUT'])
def api_update_usuario(id_usuario):
    """Update user's nombre, id_cargo, id_ciudad and/or id_jefe. Logs an audit warning when the user's city regional
    is not included in the jefe's regionals (but does not block the update).
    """
    data = request.get_json(force=True)
    nombre = data.get('nombre')
    id_cargo = data.get('id_cargo')
    id_ciudad = data.get('id_ciudad')
    id_jefe = data.get('id_jefe')
    linea_id = data.get('linea_id')
    if linea_id is not None:
        try:
            linea_id = int(linea_id)
        except Exception:
            linea_id = None

    if nombre is not None:
        nombre = str(nombre).strip()

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_update_usuario: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # Verify user exists
    cur.execute("SELECT id_usuario, nombre, id_ciudad, id_jefe FROM usuarios WHERE id_usuario = %s", (id_usuario,))
    existing = cur.fetchone()
    if not existing:
        cur.close(); db.close()
        return make_response(jsonify({"error": "Usuario no existe."}), 404)

    # Validate ciudad and jefe if provided
    try:
        nombre_ciudad = None
        id_regional_ciudad = None
        if id_ciudad is not None:
            cur.execute("SELECT id_regional, nombre_ciudad FROM ciudades WHERE id_ciudad = %s", (id_ciudad,))
            crow = cur.fetchone()
            if not crow:
                cur.close(); db.close()
                return make_response(jsonify({"error": "Ciudad no existe."}), 400)
            id_regional_ciudad = crow.get('id_regional')
            nombre_ciudad = crow.get('nombre_ciudad')

        nombre_jefe = None
        jefe_regionals = []
        if id_jefe is not None:
            # prefer many-to-many table
            cur.execute("SELECT GROUP_CONCAT(id_regional) AS regs FROM jefe_regional WHERE id_jefe = %s", (id_jefe,))
            jr = cur.fetchone()
            if jr and jr.get('regs'):
                jefe_regionals = [int(x) for x in str(jr['regs']).split(',') if x]
            else:
                # fallback to legacy column
                cur.execute("SELECT id_regional, nombre_jefe FROM jefes WHERE id_jefe = %s", (id_jefe,))
                jrow = cur.fetchone()
                if not jrow:
                    cur.close(); db.close()
                    return make_response(jsonify({"error": "Jefe no existe."}), 400)
                if jrow.get('id_regional'):
                    jefe_regionals = [int(jrow.get('id_regional'))]
                nombre_jefe = jrow.get('nombre_jefe')

        # If both city and jefe provided (or present after update), check for regional mismatch and log WARNING
        # Determine city regional to check: prefer new id_ciudad, otherwise existing
        check_city_reg = id_regional_ciudad if id_regional_ciudad is not None else existing.get('id_ciudad')
        if id_ciudad is None and existing.get('id_ciudad'):
            # get regional for existing city
            cur.execute("SELECT id_regional, nombre_ciudad FROM ciudades WHERE id_ciudad = %s", (existing.get('id_ciudad'),))
            crow2 = cur.fetchone()
            if crow2:
                check_city_reg = crow2.get('id_regional')
                nombre_ciudad = crow2.get('nombre_ciudad')

        # Determine jefe regionals to check: prefer new id_jefe, else existing
        check_jefe = id_jefe if id_jefe is not None else existing.get('id_jefe')
        jefe_regs_to_check = jefe_regionals
        if not jefe_regs_to_check and check_jefe:
            cur.execute("SELECT GROUP_CONCAT(id_regional) AS regs FROM jefe_regional WHERE id_jefe = %s", (check_jefe,))
            jr2 = cur.fetchone()
            if jr2 and jr2.get('regs'):
                jefe_regs_to_check = [int(x) for x in str(jr2['regs']).split(',') if x]
            else:
                cur.execute("SELECT id_regional, nombre_jefe FROM jefes WHERE id_jefe = %s", (check_jefe,))
                jrow2 = cur.fetchone()
                if jrow2 and jrow2.get('id_regional'):
                    jefe_regs_to_check = [int(jrow2.get('id_regional'))]
                    nombre_jefe = nombre_jefe or jrow2.get('nombre_jefe')

        try:
            if check_city_reg is not None and jefe_regs_to_check is not None and check_jefe is not None:
                # if city regional not in jefe's region list -> warn
                if int(check_city_reg) not in jefe_regs_to_check:
                    # get usuario nombre for log
                    usuario_nombre = nombre or existing.get('nombre')
                    app.logger.warning(f"[audit] jefe={check_jefe}({nombre_jefe}, regs={jefe_regs_to_check}) usuario={id_usuario}({usuario_nombre}, reg={check_city_reg}) accion=update_usuario detalle=\"regional_mismatch_permitido\"")
        except Exception:
            app.logger.exception('audit log failed on api_update_usuario')

    except Exception as e:
        print(f"api_update_usuario: validation error: {e}")
        traceback.print_exc()
        cur.close(); db.close()
        return make_response(jsonify({"error": "Error interno en validación."}), 500)

    # Build update statement
    updates = []
    params = []
    if nombre is not None:
        updates.append("nombre = %s"); params.append(nombre)
    if id_cargo is not None:
        updates.append("id_cargo = %s"); params.append(id_cargo)
    if id_ciudad is not None:
        updates.append("id_ciudad = %s"); params.append(id_ciudad)
    if id_jefe is not None:
        updates.append("id_jefe = %s"); params.append(id_jefe)

    if updates:
        sql = "UPDATE usuarios SET " + ", ".join(updates) + " WHERE id_usuario = %s"
        params.append(id_usuario)
        try:
            cur.execute(sql, params)
            db.commit()
            # Touch lines for this usuario so they surface as recently-modified
            try:
                if has_fecha_modificacion():
                    try:
                        cur.execute("UPDATE lineas SET fecha_modificacion = NOW() WHERE id_usuario = %s", (id_usuario,))
                        db.commit()
                    except Exception:
                        db.rollback()
                        app.logger.exception('Could not touch lineas after usuario update')
                # touch in-memory for this user's lines
                try:
                    touch_lineas_for_usuario(db, id_usuario)
                except Exception:
                    app.logger.exception('Could not touch in memory after usuario update')
            except Exception:
                pass
            # If caller requested, update specific linea's ciudad/regional to keep detalle page consistent
            try:
                if linea_id and id_ciudad is not None:
                    try:
                        if has_fecha_modificacion():
                            cur.execute("UPDATE lineas SET id_ciudad = %s, fecha_modificacion = NOW() WHERE id_linea = %s", (id_ciudad, int(linea_id)))
                        else:
                            cur.execute("UPDATE lineas SET id_ciudad = %s WHERE id_linea = %s", (id_ciudad, int(linea_id)))
                        db.commit()
                    except Exception:
                        db.rollback()
                        app.logger.exception('Could not update linea ciudad after usuario update')
            except Exception:
                app.logger.exception('linea update check failed')
            # If no specific linea_id was provided, but the user's ciudad changed, update all lines for that user
            try:
                if (linea_id is None) and (id_ciudad is not None):
                    try:
                        if has_fecha_modificacion():
                            cur.execute("UPDATE lineas SET id_ciudad = %s, fecha_modificacion = NOW() WHERE id_usuario = %s", (id_ciudad, id_usuario))
                        else:
                            cur.execute("UPDATE lineas SET id_ciudad = %s WHERE id_usuario = %s", (id_ciudad, id_usuario))
                        db.commit()
                    except Exception:
                        db.rollback()
                        app.logger.exception('Could not update lineas ciudad after usuario update')
            except Exception:
                app.logger.exception('bulk linea update check failed')
        except mysql.connector.Error as e:
            db.rollback()
            print(f"api_update_usuario: DB update error: {e}")
            traceback.print_exc()
            cur.close(); db.close()
            return make_response(jsonify({"error": "Error interno al actualizar usuario."}), 500)

    # Return updated user
    cur.execute("SELECT id_usuario, nombre, id_cargo, id_ciudad, id_jefe FROM usuarios WHERE id_usuario = %s", (id_usuario,))
    updated = cur.fetchone()
    cur.close(); db.close()
    return make_response(jsonify({
        "id_usuario": updated.get('id_usuario'),
        "nombre": updated.get('nombre'),
        "id_cargo": updated.get('id_cargo'),
        "id_ciudad": updated.get('id_ciudad'),
        "id_jefe": updated.get('id_jefe')
    }), 200)


# =========================
# API: Crear línea
# =========================
@app.route('/api/lineas', methods=['POST'])
def api_create_linea():
    data = request.get_json(force=True)
    numero_linea = (data.get('numero_linea') or '').strip()
    id_usuario = data.get('id_usuario')
    id_plan = data.get('id_plan')
    id_tipo_sim = data.get('id_tipo_sim')
    id_estado = data.get('id_estado')
    id_ciudad = data.get('id_ciudad')
    id_jefe = data.get('id_jefe')
    observacion = data.get('observacion')

    if not numero_linea or not id_usuario:
        return make_response(jsonify({"error": "Campos obligatorios faltantes."}), 400)

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_get_planes: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # Verificar duplicado de número
    cur.execute("SELECT COUNT(1) AS cnt FROM lineas WHERE numero_linea = %s", (numero_linea,))
    if cur.fetchone()['cnt'] > 0:
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Número de línea ya existe."}), 409)

    # Verificar usuario existe
    cur.execute("SELECT COUNT(1) AS cnt FROM usuarios WHERE id_usuario = %s", (id_usuario,))
    if cur.fetchone()['cnt'] == 0:
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Usuario no existe."}), 400)

    # Insertar con transacción
    try:
        db.start_transaction()
        if has_fecha_modificacion():
            cur.execute("INSERT INTO lineas (numero_linea, id_usuario, id_plan, id_tipo_sim, id_estado, id_ciudad, id_jefe, observacion, fecha_modificacion) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())",
                        (numero_linea, id_usuario, id_plan, id_tipo_sim, id_estado, id_ciudad, id_jefe, observacion))
        else:
            cur.execute("INSERT INTO lineas (numero_linea, id_usuario, id_plan, id_tipo_sim, id_estado, id_ciudad, id_jefe, observacion) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                        (numero_linea, id_usuario, id_plan, id_tipo_sim, id_estado, id_ciudad, id_jefe, observacion))
        new_id = cur.lastrowid
        db.commit()
    except mysql.connector.Error as e:
        db.rollback()
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Error interno"}), 500)

    # Touch in-memory so the newly created linea surfaces first in listings
    try:
        touch_linea_in_memory(new_id)
    except Exception:
        app.logger.exception('Could not touch linea in memory after creating linea')

    # Registrar novedad inicial de creación/asignación
    try:
        cur.execute("SELECT nombre FROM usuarios WHERE id_usuario = %s LIMIT 1", (id_usuario,))
        usuario_row = cur.fetchone()
        nombre_usuario = usuario_row['nombre'] if usuario_row else '—'
        registrar_novedad(
            db=db,
            id_linea=new_id,
            tipo='ASIGNACION',
            detalle=f'Línea creada y asignada a {nombre_usuario}',
            valor_anterior='—',
            valor_nuevo=nombre_usuario
        )
    except Exception:
        app.logger.exception('Could not register initial assignment novedad')

    cur.execute("SELECT id_linea, numero_linea FROM lineas WHERE id_linea = %s", (new_id,))
    nl = cur.fetchone()
    cur.close()
    db.close()
    return make_response(jsonify({"id_linea": nl['id_linea'], "numero_linea": nl['numero_linea']}), 201)


@app.route('/api/lineas/check_numero', methods=['GET'])
def api_check_numero():
    """Check whether a given numero_linea exists. Returns {exists: bool, id_linea: int|null}."""
    numero = (request.args.get('numero') or '').strip()
    if not numero:
        return make_response(jsonify({"error": "Parametro 'numero' requerido."}), 400)
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT id_linea FROM lineas WHERE numero_linea = %s LIMIT 1", (numero,))
        row = cur.fetchone()
        cur.close()
        db.close()
        if row:
            return jsonify({"exists": True, "id_linea": row.get('id_linea')})
        else:
            return jsonify({"exists": False, "id_linea": None})
    except Exception as e:
        app.logger.exception('api_check_numero failed')
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass
        return make_response(jsonify({"error": "Error interno al consultar."}), 500)


# =========================
# API: Planes (CRUD mínimo: GET, POST)
# =========================
@app.route('/api/planes', methods=['GET'])
def api_get_planes():
    search = (request.args.get('search') or '').strip()
    limit = request.args.get('limit', type=int) or 50
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_create_plan: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)
    sql = "SELECT id_plan, nombre_plan, gigas, precio_mensual FROM planes"
    params = []
    if search:
        sql += " WHERE LOWER(nombre_plan) LIKE %s"
        params.append(f"%{search.lower()}%")
    sql += " ORDER BY nombre_plan LIMIT %s"
    params.append(limit)
    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    db.close()
    return jsonify([{"id_plan": r['id_plan'], "nombre_plan": r['nombre_plan'], "gigas": r['gigas'], "precio_mensual": r['precio_mensual']} for r in rows])


@app.route('/api/planes', methods=['POST'])
def api_create_plan():
    data = request.get_json(force=True)
    nombre = (data.get('nombre_plan') or '').strip()
    gigas = data.get('gigas')
    precio = data.get('precio_mensual')
    if not nombre:
        return make_response(jsonify({"error": "Nombre de plan obligatorio."}), 400)
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_get_jefes: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)
    # unicidad
    cur.execute("SELECT COUNT(1) AS cnt FROM planes WHERE nombre_plan = %s", (nombre,))
    if cur.fetchone()['cnt'] > 0:
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Plan ya existe."}), 409)
    try:
        cur.execute("INSERT INTO planes (nombre_plan, gigas, precio_mensual) VALUES (%s, %s, %s)", (nombre, gigas, precio))
        new_id = cur.lastrowid
        db.commit()
    except mysql.connector.Error:
        db.rollback()
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Error interno"}), 500)
    cur.execute("SELECT id_plan, nombre_plan, gigas, precio_mensual FROM planes WHERE id_plan = %s", (new_id,))
    r = cur.fetchone()
    cur.close()
    db.close()
    return make_response(jsonify({"id_plan": r['id_plan'], "nombre_plan": r['nombre_plan'], "gigas": r['gigas'], "precio_mensual": r['precio_mensual']}), 201)


# =========================
# API: Listado (GET) - formato estable {items, total}
# =========================
@app.route('/api/lineas', methods=['GET'])
def api_get_lineas():
    # Parse pagination params
    try:
        page = int(request.args.get('page', 1))
    except Exception:
        page = 1
    try:
        per_page = int(request.args.get('per_page', 10))
    except Exception:
        per_page = 10
    if page < 1:
        page = 1
    if per_page < 1:
        per_page = 10

    q = (request.args.get('q') or '').strip()
    plan_f = (request.args.get('plan') or '').strip()
    tipo_sim_f = (request.args.get('tipo_sim') or '').strip()
    estado_f = (request.args.get('estado') or '').strip()

    where = []
    params = []
    if plan_f:
        where.append("LOWER(p.nombre_plan) = %s")
        params.append(plan_f.lower())
    if tipo_sim_f:
        where.append("LOWER(ts.nombre_tipo) = %s")
        params.append(tipo_sim_f.lower())
    if estado_f:
        # allow filtering by numeric id (from catalog select) or by name
        if estado_f.isdigit():
            where.append("es.id_estado = %s")
            params.append(int(estado_f))
        else:
            where.append("LOWER(es.nombre_estado) LIKE %s")
            params.append(f"{estado_f.lower()}%")
    if q:
        ql = q.lower()
        like = f"%{ql}%"
        where.append("(LOWER(l.numero_linea) LIKE %s OR LOWER(u.nombre) LIKE %s OR LOWER(c.nombre_cargo) LIKE %s OR LOWER(ci.nombre_ciudad) LIKE %s OR LOWER(p.nombre_plan) LIKE %s OR LOWER(ts.nombre_tipo) LIKE %s)")
        params += [like, like, like, like, like, like]

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    # Connect to DB
    try:
        db = conectar_db()
        cur = db.cursor()
    except Exception as e:
        print(f"api_get_lineas: DB connection failed: {e}")
        return make_response(jsonify({"items": [], "total": 0, "warning": "Base de datos no disponible."}), 200)

    # COUNT with timing
    try:
        start = time.time()
        cur.execute(f"""
            SELECT COUNT(*)
            FROM lineas l
            LEFT JOIN usuarios      u  ON u.id_usuario   = l.id_usuario
            LEFT JOIN cargos        c  ON c.id_cargo     = u.id_cargo
            LEFT JOIN ciudades      ci ON ci.id_ciudad   = l.id_ciudad
            LEFT JOIN regionales    r  ON r.id_regional  = ci.id_regional
            LEFT JOIN tipos_sim     ts ON ts.id_tipo_sim = l.id_tipo_sim
            LEFT JOIN estados_linea es ON es.id_estado   = l.id_estado
            LEFT JOIN planes        p  ON p.id_plan      = l.id_plan
            {where_sql}
        """, params)
        duration = time.time() - start
    except Exception as e:
        try:
            cur.close()
            db.close()
        except Exception:
            pass
        print(f"[ERROR] api_get_lineas COUNT failed: {e}")
        return make_response(jsonify({"items": [], "total": 0, "warning": "Error interno en conteo"}), 200)

    print(f"[DEBUG] api_get_lineas COUNT duration: {duration:.2f}s")
    if duration > 2.0:
        try:
            cur.close()
            db.close()
        except Exception:
            pass
        return make_response(jsonify({"items": [], "total": 0, "warning": "Consulta demasiado lenta; filtra la búsqueda."}), 200)

    try:
        total = cur.fetchone()[0]
    except Exception:
        total = 0
    cur.close()

    if total > 20000:
        try:
            db.close()
        except Exception:
            pass
        return make_response(jsonify({"items": [], "total": total, "warning": "Demasiados registros para cargar de una vez."}), 200)

    # Fetch matching rows
    try:
        # Paginar y ordenar en SQL para persistencia y rendimiento
        offset = (page - 1) * per_page
        cur = db.cursor(dictionary=True)
        start = time.time()
        has_fmod = has_fecha_modificacion()
        fecha_field = "l.fecha_modificacion AS fecha_modificacion," if has_fmod else ""

        order_clause = "ORDER BY l.fecha_modificacion DESC, l.id_linea DESC" if has_fmod else "ORDER BY l.id_linea DESC"

        sql = f"""
            SELECT 
                l.id_linea                 AS id,
                l.numero_linea             AS numero,
                u.nombre                   AS usuario,
                c.nombre_cargo             AS cargo,
                ci.nombre_ciudad           AS ciudad,
                r.nombre_regional          AS regional,
                p.nombre_plan              AS plan,
                {fecha_field}
                CASE WHEN p.gigas IS NULL THEN 'Ilimitadas' ELSE CONCAT(p.gigas, ' GB') END AS gigas_plan,
                ts.nombre_tipo             AS tipo_sim,
                es.nombre_estado           AS estado,
                COALESCE(l.observacion, '') AS observacion
            FROM lineas l
            LEFT JOIN usuarios      u  ON u.id_usuario   = l.id_usuario
            LEFT JOIN cargos        c  ON c.id_cargo     = u.id_cargo
            LEFT JOIN ciudades      ci ON ci.id_ciudad   = l.id_ciudad
            LEFT JOIN regionales    r  ON r.id_regional  = ci.id_regional
            LEFT JOIN tipos_sim     ts ON ts.id_tipo_sim = l.id_tipo_sim
            LEFT JOIN estados_linea es ON es.id_estado   = l.id_estado
            LEFT JOIN planes        p  ON p.id_plan      = l.id_plan
            {where_sql}
            {order_clause}
            LIMIT %s OFFSET %s
        """

        params2 = params + [per_page, offset]
        cur.execute(sql, params2)
        duration = time.time() - start
    except Exception as e:
        try:
            cur.close()
            db.close()
        except Exception:
            pass
        print(f"[ERROR] api_get_lineas SELECT failed: {e}")
        return make_response(jsonify({"items": [], "total": total, "warning": "Error interno en consulta"}), 200)

    print(f"[DEBUG] api_get_lineas SELECT duration: {duration:.2f}s")
    if duration > 2.0:
        try:
            cur.close()
            db.close()
        except Exception:
            pass
        return make_response(jsonify({"items": [], "total": total, "warning": "Consulta paginada demasiado lenta."}), 200)

    try:
        rows = cur.fetchall()
    finally:
        try:
            cur.close()
            db.close()
        except Exception:
            pass

    items = []
    for r in rows:
        items.append({
            "id": r['id'],
            "numero": r['numero'],
            "usuario": r['usuario'] or None,
            "cargo": r['cargo'] or None,
            "ciudad": r['ciudad'] or None,
            "regional": r.get('regional'),
            "plan": r['plan'] or None,
            "gigas_plan": r.get('gigas_plan'),
            "tipo_sim": r['tipo_sim'] or None,
            "estado": r['estado'] or None,
            "observacion": r['observacion'] or None,
        })

    return jsonify({"items": items, "total": total})


# =========================
# API: Catálogos (cargados por AJAX)
# =========================
@app.route('/api/catalogos/planes', methods=['GET'])
def api_catalogo_planes():
    try:
        db = conectar_db()
        cur = db.cursor()
    except Exception as e:
        print(f"api_catalogo_planes: DB connection failed: {e}")
        return make_response(jsonify({"error": "DB unavailable"}), 503)
    try:
        cur.execute("SELECT DISTINCT nombre_plan FROM planes ORDER BY nombre_plan")
        rows = [r[0] for r in cur.fetchall()]
    except Exception as e:
        cur.close()
        db.close()
        print(f"api_catalogo_planes error: {e}")
        return make_response(jsonify({"error": "Query failed"}), 500)
    cur.close()
    db.close()
    return jsonify(rows)


@app.route('/api/catalogos/tipos_sim', methods=['GET'])
def api_catalogo_tipos():
    try:
        db = conectar_db()
        cur = db.cursor()
    except Exception as e:
        print(f"api_catalogo_tipos: DB connection failed: {e}")
        return make_response(jsonify({"error": "DB unavailable"}), 503)
    try:
        cur.execute("SELECT DISTINCT nombre_tipo FROM tipos_sim ORDER BY nombre_tipo")
        rows = [r[0] for r in cur.fetchall()]
    except Exception as e:
        cur.close()
        db.close()
        print(f"api_catalogo_tipos error: {e}")
        return make_response(jsonify({"error": "Query failed"}), 500)
    cur.close()
    db.close()
    return jsonify(rows)


# Exportar líneas filtradas como CSV (Excel compatible)
@app.route('/lineas/export', methods=['GET'])
def export_lineas():
    # aceptar mismos filtros que /api/lineas pero sin paginación
    q = (request.args.get('q') or '').strip()
    plan_f = (request.args.get('plan') or '').strip()
    tipo_sim_f = (request.args.get('tipo_sim') or '').strip()
    estado_f = (request.args.get('estado') or '').strip()

    where = []
    params = []
    if plan_f:
        where.append("LOWER(p.nombre_plan) = %s")
        params.append(plan_f.lower())
    if tipo_sim_f:
        where.append("LOWER(ts.nombre_tipo) = %s")
        params.append(tipo_sim_f.lower())
    if estado_f:
        if estado_f.isdigit():
            where.append("es.id_estado = %s")
            params.append(int(estado_f))
        else:
            where.append("LOWER(es.nombre_estado) LIKE %s")
            params.append(f"{estado_f.lower()}%")
    if q:
        ql = q.lower()
        like = f"%{ql}%"
        where.append("(LOWER(l.numero_linea) LIKE %s OR LOWER(u.nombre) LIKE %s OR LOWER(c.nombre_cargo) LIKE %s OR LOWER(ci.nombre_ciudad) LIKE %s OR LOWER(p.nombre_plan) LIKE %s OR LOWER(ts.nombre_tipo) LIKE %s)")
        params += [like, like, like, like, like, like]

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"export_lineas: DB connection failed: {e}")
        return make_response('DB unavailable', 503)

    try:
        sql = f"""
            SELECT
                l.id_linea AS id,
                l.numero_linea AS numero,
                u.nombre AS usuario,
                c.nombre_cargo AS cargo,
                ci.nombre_ciudad AS ciudad,
                r.nombre_regional AS regional,
                p.nombre_plan AS plan,
                CASE WHEN p.gigas IS NULL THEN 'Ilimitadas' ELSE CONCAT(p.gigas, ' GB') END AS gigas_plan,
                ts.nombre_tipo AS tipo_sim,
                es.nombre_estado AS estado,
                COALESCE(l.observacion, '') AS observacion
            FROM lineas l
            LEFT JOIN usuarios      u  ON u.id_usuario   = l.id_usuario
            LEFT JOIN cargos        c  ON c.id_cargo     = u.id_cargo
            LEFT JOIN ciudades      ci ON ci.id_ciudad   = l.id_ciudad
            LEFT JOIN regionales    r  ON r.id_regional  = ci.id_regional
            LEFT JOIN tipos_sim     ts ON ts.id_tipo_sim = l.id_tipo_sim
            LEFT JOIN estados_linea es ON es.id_estado   = l.id_estado
            LEFT JOIN planes        p  ON p.id_plan      = l.id_plan
            {where_sql}
            ORDER BY l.id_linea DESC
        """
        cur.execute(sql, params)
        rows = cur.fetchall()
    except Exception as e:
        try:
            cur.close(); db.close()
        except Exception:
            pass
        print(f"export_lineas SELECT failed: {e}")
        return make_response('Query failed', 500)

    # Construir CSV en memoria (UTF-8 con BOM para Excel)
    output = io.StringIO()
    writer = csv.writer(output)
    header = ['ID','Número','Usuario','Cargo','Ciudad','Regional','Plan','Gigas Plan','Tipo SIM','Estado','Observación']
    writer.writerow(header)
    for r in rows:
        writer.writerow([
            r.get('id'), r.get('numero'), r.get('usuario'), r.get('cargo'), r.get('ciudad'), r.get('regional'), r.get('plan'), r.get('gigas_plan'), r.get('tipo_sim'), r.get('estado'), r.get('observacion')
        ])

    csv_data = output.getvalue()
    output.close()
    try:
        cur.close(); db.close()
    except Exception:
        pass

    # Prepend BOM so Excel detects UTF-8 properly
    bom = '\ufeff'
    resp = make_response(bom + csv_data)
    resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
    resp.headers['Content-Disposition'] = 'attachment; filename=lineas_export.csv'
    return resp


# Exportar líneas filtradas como XLSX (mejor experiencia en Excel)
@app.route('/lineas/export.xlsx', methods=['GET'])
def export_lineas_xlsx():
    if not OPENPYXL_AVAILABLE:
        return make_response('openpyxl no está instalado. Ejecuta: pip install openpyxl', 500)

    # mismos filtros que la ruta CSV
    q = (request.args.get('q') or '').strip()
    plan_f = (request.args.get('plan') or '').strip()
    tipo_sim_f = (request.args.get('tipo_sim') or '').strip()
    estado_f = (request.args.get('estado') or '').strip()

    where = []
    params = []
    if plan_f:
        where.append("LOWER(p.nombre_plan) = %s")
        params.append(plan_f.lower())
    if tipo_sim_f:
        where.append("LOWER(ts.nombre_tipo) = %s")
        params.append(tipo_sim_f.lower())
    if estado_f:
        if estado_f.isdigit():
            where.append("es.id_estado = %s")
            params.append(int(estado_f))
        else:
            where.append("LOWER(es.nombre_estado) LIKE %s")
            params.append(f"{estado_f.lower()}%")
    if q:
        ql = q.lower()
        like = f"%{ql}%"
        where.append("(LOWER(l.numero_linea) LIKE %s OR LOWER(u.nombre) LIKE %s OR LOWER(c.nombre_cargo) LIKE %s OR LOWER(ci.nombre_ciudad) LIKE %s OR LOWER(p.nombre_plan) LIKE %s OR LOWER(ts.nombre_tipo) LIKE %s)")
        params += [like, like, like, like, like, like]

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"export_lineas_xlsx: DB connection failed: {e}")
        return make_response('DB unavailable', 503)

    try:
        sql = f"""
            SELECT
                l.id_linea AS id,
                l.numero_linea AS numero,
                u.nombre AS usuario,
                c.nombre_cargo AS cargo,
                ci.nombre_ciudad AS ciudad,
                r.nombre_regional AS regional,
                p.nombre_plan AS plan,
                CASE WHEN p.gigas IS NULL THEN 'Ilimitadas' ELSE CONCAT(p.gigas, ' GB') END AS gigas_plan,
                ts.nombre_tipo AS tipo_sim,
                es.nombre_estado AS estado,
                COALESCE(l.observacion, '') AS observacion
            FROM lineas l
            LEFT JOIN usuarios      u  ON u.id_usuario   = l.id_usuario
            LEFT JOIN cargos        c  ON c.id_cargo     = u.id_cargo
            LEFT JOIN ciudades      ci ON ci.id_ciudad   = l.id_ciudad
            LEFT JOIN regionales    r  ON r.id_regional  = ci.id_regional
            LEFT JOIN tipos_sim     ts ON ts.id_tipo_sim = l.id_tipo_sim
            LEFT JOIN estados_linea es ON es.id_estado   = l.id_estado
            LEFT JOIN planes        p  ON p.id_plan      = l.id_plan
            {where_sql}
            ORDER BY l.id_linea DESC
        """
        cur.execute(sql, params)
        rows = cur.fetchall()
    except Exception as e:
        try:
            cur.close(); db.close()
        except Exception:
            pass
        print(f"export_lineas_xlsx SELECT failed: {e}")
        return make_response('Query failed', 500)

    # Crear workbook y aplicar formato más amigable para Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Lineas'

    headers = ['ID','Número','Usuario','Cargo','Ciudad','Regional','Plan','Gigas Plan','Tipo SIM','Estado','Observación']

    # Título en la primera fila (merge sobre todas las columnas)
    title = f"Exportación de líneas - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Fila de información (filtros aplicados) en la segunda fila si aplica
    info_items = []
    if q:
        info_items.append(f"q={q}")
    if plan_f:
        info_items.append(f"plan={plan_f}")
    if tipo_sim_f:
        info_items.append(f"tipo_sim={tipo_sim_f}")
    if estado_f:
        info_items.append(f"estado={estado_f}")
    info_text = ' | '.join(info_items)
    header_row_idx = 2
    if info_text:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        info_cell = ws.cell(row=2, column=1, value=info_text)
        info_cell.alignment = Alignment(horizontal='center')
        header_row_idx = 3
    else:
        header_row_idx = 2

    # Escribir fila de encabezados con estilo
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin = Side(border_style="thin", color="000000")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Agregar filas de datos
    for r in rows:
        ws.append([
            r.get('id'), r.get('numero'), r.get('usuario'), r.get('cargo'), r.get('ciudad'), r.get('regional'), r.get('plan'), r.get('gigas_plan'), r.get('tipo_sim'), r.get('estado'), r.get('observacion')
        ])

    # Aplicar color alternado a las filas de datos
    data_start = header_row_idx + 1
    light_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    for i, row in enumerate(ws.iter_rows(min_row=data_start, max_row=ws.max_row, min_col=1, max_col=len(headers)), start=0):
        fill = light_fill if i % 2 == 0 else alt_fill
        for cell in row:
            # preservar encabezado/other formatting
            if cell.row >= data_start:
                cell.fill = fill

    # Congelar paneles justo debajo de la fila de encabezado
    ws.freeze_panes = ws.cell(row=data_start, column=1)

    # Ajustar anchos de columna básicos
    for i, col in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col:
            try:
                val = str(cell.value or '')
            except Exception:
                val = ''
            if len(val) > max_len:
                max_len = len(val)
        adjusted_width = min(max(10, max_len + 2), 60)
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    # Guardar en memoria
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    try:
        cur.close(); db.close()
    except Exception:
        pass

    resp = make_response(bio.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename=lineas_export.xlsx'
    return resp


@app.route('/api/catalogos/estados', methods=['GET'])
def api_catalogo_estados():
    try:
        db = conectar_db()
        cur = db.cursor()
    except Exception as e:
        print(f"api_catalogo_estados: DB connection failed: {e}")
        return make_response(jsonify({"error": "DB unavailable"}), 503)
    try:
        cur.execute("SELECT id_estado, nombre_estado FROM estados_linea ORDER BY nombre_estado")
        rows = [
            {"id": r[0], "nombre": r[1]}
            for r in cur.fetchall()
            if (r[1] or '').strip().lower() not in ('cesionada', 'liberada')
        ]
    except Exception as e:
        cur.close()
        db.close()
        print(f"api_catalogo_estados error: {e}")
        return make_response(jsonify({"error": "Query failed"}), 500)
    cur.close()
    db.close()
    return jsonify(rows)


@app.route('/api/estados/ensure_cesion', methods=['POST'])
def api_ensure_estados_cesion():
    """Ensure the states for cesión exist and unify 'Transferida' into 'Cesionada'.
    Behavior:
    - Ensure 'En cesión' exists (for marking a pending transfer; note: not applied automatically anywhere).
    - Ensure 'Cesionada' exists as the canonical 'transferida' final state.
    - If a state named 'Transferida' exists but 'Cesionada' does not, rename 'Transferida' -> 'Cesionada' to unify names.
    Returns which names were created or renamed.
    """
    wanted = ["En cesión", "Cesionada"]
    created = []
    renamed = []
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)

        #  'Transferida' 
        cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", ("transferida",))
        tf = cur.fetchone()

        # Ensure wanted names exist (create if missing). Do NOT rename any existing rows.
        for name in wanted:
            cur.execute("SELECT id_estado FROM estados_linea WHERE LOWER(nombre_estado) = %s LIMIT 1", (name.lower(),))
            if not cur.fetchone():
                try:
                    cur.execute("INSERT INTO estados_linea (nombre_estado) VALUES (%s)", (name,))
                    db.commit()
                    created.append(name)
                except Exception:
                    db.rollback()

        cur.close()
        db.close()
        return jsonify({"ok": True, "created": created, "transferida_exists": bool(tf), "wanted": wanted})
    except Exception as e:
        app.logger.exception('api_ensure_estados_cesion failed')
        try:
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass
        return make_response(jsonify({"error": "Error interno al asegurar estados."}), 500)


# -----------------
# Debug endpoints (temporary)
# -----------------
@app.route('/debug/touch_linea/<int:linea_id>', methods=['POST','GET'])
def debug_touch_linea(linea_id):
    try:
        touch_linea_in_memory(linea_id)
        return jsonify({"ok": True, "id_linea": linea_id, "touched_at": app.config.get('LINEAS_TOUCH_CACHE', {}).get(int(linea_id)).isoformat()})
    except Exception as e:
        app.logger.exception('debug_touch_linea failed')
        return make_response(jsonify({"error": "failed"}), 500)


@app.route('/debug/list_touches', methods=['GET'])
def debug_list_touches():
    try:
        cache = app.config.get('LINEAS_TOUCH_CACHE', {})
        out = {str(k): v.isoformat() for k, v in cache.items()}
        return jsonify(out)
    except Exception as e:
        app.logger.exception('debug_list_touches failed')
        return make_response(jsonify({"error": "failed"}), 500)


@app.route('/debug/clear_touches', methods=['POST','GET'])
def debug_clear_touches():
    try:
        app.config['LINEAS_TOUCH_CACHE'] = {}
        return jsonify({"ok": True, "cleared": True})
    except Exception:
        app.logger.exception('debug_clear_touches failed')
        return make_response(jsonify({"error": "failed"}), 500)


@app.route('/debug/remove_touch/<int:linea_id>', methods=['POST','GET'])
def debug_remove_touch(linea_id):
    try:
        cache = app.config.get('LINEAS_TOUCH_CACHE', {})
        removed = cache.pop(int(linea_id), None) is not None
        app.config['LINEAS_TOUCH_CACHE'] = cache
        return jsonify({"ok": True, "removed": removed})
    except Exception:
        app.logger.exception('debug_remove_touch failed')
        return make_response(jsonify({"error": "failed"}), 500)


@app.route('/debug/db_status', methods=['GET'])
def debug_db_status():
    """Endpoint de depuración: intenta conectar a la BD y ejecutar una consulta mínima.
    Útil para comprobar si la aplicación puede acceder a la base de datos tras reiniciar el servicio.
    """
    try:
        db = conectar_db()
        cur = db.cursor()
        cur.execute("SELECT 1")
        cur.close()
        db.close()
        return jsonify({"ok": True, "msg": "DB reachable"})
    except Exception as e:
        try:
            # intentar cerrar si existe
            cur.close()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass
        app.logger.exception('debug_db_status failed')
        return make_response(jsonify({"ok": False, "error": str(e)}), 500)


@app.route('/debug/mock_lineas', methods=['GET'])
def debug_mock_lineas():
    """Devuelve un JSON de ejemplo para comprobar que la UI y JS renderizan correctamente."""
    items = [
        {"id": 1, "numero": "3100000001", "usuario": "Juan Perez", "cargo": "Analista", "ciudad": "Bogotá", "plan": "Plan A", "tipo_sim": "Prepago", "estado": "Activa", "observacion": "Ejemplo"},
        {"id": 2, "numero": "3100000002", "usuario": "María Ruiz", "cargo": "Gerente", "ciudad": "Medellín", "plan": "Plan B", "tipo_sim": "Pospago", "estado": "Suspendida", "observacion": "Ejemplo 2"}
    ]
    return jsonify({"items": items, "total": len(items)})


# =========================
# Reportes: Historial de novedades
# =========================
@app.route('/reportes')
def reportes_index():
    # Mostrar estadísticas básicas: total de líneas, por tipo SIM y por estado
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        app.logger.exception('reportes_index: DB connection failed')
        return render_template('reportes.html', stats=None, by_tipo=[], by_estado=[], by_plan=[])

    # Total líneas
    cur.execute("SELECT COUNT(1) AS total FROM lineas")
    total = cur.fetchone().get('total', 0)

    # Por tipo SIM
    cur.execute("SELECT COALESCE(ts.nombre_tipo, 'Sin asignar') AS nombre, COUNT(1) AS cnt FROM lineas l LEFT JOIN tipos_sim ts ON ts.id_tipo_sim = l.id_tipo_sim GROUP BY nombre ORDER BY cnt DESC")
    by_tipo = cur.fetchall()

    # Por estado (agrupando 'Liberada' como 'Inactiva' y 'Cesionada' como 'Activo')
    cur.execute("""
        SELECT 
            CASE 
                WHEN LOWER(es.nombre_estado) = 'liberada' THEN 'Inactiva'
                WHEN LOWER(es.nombre_estado) = 'cesionada' THEN 'Activo'
                ELSE COALESCE(es.nombre_estado, 'Sin asignar')
            END AS nombre,
            COUNT(1) AS cnt 
        FROM lineas l 
        LEFT JOIN estados_linea es ON es.id_estado = l.id_estado 
        GROUP BY nombre 
        ORDER BY cnt DESC
    """)
    by_estado = cur.fetchall()

    # Por plan (top 10)
    cur.execute("SELECT COALESCE(p.nombre_plan, 'Sin asignar') AS nombre, COUNT(1) AS cnt FROM lineas l LEFT JOIN planes p ON p.id_plan = l.id_plan GROUP BY nombre ORDER BY cnt DESC LIMIT 10")
    by_plan = cur.fetchall()

    cur.close()
    db.close()

    stats = { 'total_lineas': total }
    return render_template('reportes.html', stats=stats, by_tipo=by_tipo, by_estado=by_estado, by_plan=by_plan)


@app.route('/reportes/novedades', methods=['GET'])
def reportes_novedades():
    """Devuelve novedades en JSON o CSV. Parámetros opcionales:
    - desde (YYYY-MM-DD)
    - hasta (YYYY-MM-DD)
    - tipo
    - id_linea
    - usuario_actor
    - format=csv|json (json por defecto)
    """
    fmt = (request.args.get('format') or 'json').lower()
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')
    tipo = (request.args.get('tipo') or '').strip()
    id_linea = request.args.get('id_linea', type=int)
    usuario_actor = (request.args.get('usuario_actor') or '').strip()

    where = []
    params = []
    if desde:
        where.append("fecha >= %s")
        params.append(desde + ' 00:00:00')
    if hasta:
        where.append("fecha <= %s")
        params.append(hasta + ' 23:59:59')
    if tipo:
        where.append("tipo = %s")
        params.append(tipo)
    if id_linea:
        where.append("id_linea = %s")
        params.append(id_linea)
    if usuario_actor:
        where.append("usuario_actor LIKE %s")
        params.append(f"%{usuario_actor}%")

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception:
        return make_response(jsonify({"error": "DB connection failed"}), 503)

    sql = f"SELECT id_novedad, id_linea, tipo, detalle, valor_anterior, valor_nuevo, usuario_actor, fecha FROM novedades_linea {where_sql} ORDER BY fecha DESC LIMIT 2000"
    try:
        cur.execute(sql, params)
        rows = cur.fetchall()
    except Exception:
        try:
            cur.close(); db.close()
        except Exception:
            pass
        return make_response(jsonify({"error": "Query failed"}), 500)

    out = []
    for r in rows:
        out.append({
            "id_novedad": r.get('id_novedad'),
            "id_linea": r.get('id_linea'),
            "tipo": r.get('tipo'),
            "detalle": r.get('detalle'),
            "valor_anterior": r.get('valor_anterior'),
            "valor_nuevo": r.get('valor_nuevo'),
            "usuario_actor": r.get('usuario_actor'),
            "fecha": r.get('fecha').isoformat() if isinstance(r.get('fecha'), datetime) else str(r.get('fecha'))
        })

    cur.close(); db.close()

    if fmt == 'csv':
        si = io.StringIO()
        writer = csv.writer(si)
        writer.writerow(['id_novedad','id_linea','tipo','detalle','valor_anterior','valor_nuevo','usuario_actor','fecha'])
        for r in out:
            writer.writerow([r['id_novedad'], r['id_linea'], r['tipo'], r['detalle'], r['valor_anterior'], r['valor_nuevo'], r['usuario_actor'], r['fecha']])
        output = make_response(si.getvalue())
        output.headers['Content-Type'] = 'text/csv; charset=utf-8'
        output.headers['Content-Disposition'] = 'attachment; filename=novedades.csv'
        return output

    return jsonify({"items": out, "total": len(out)})


# =========================
# API: Jefes (GET, POST)
# =========================
@app.route('/api/jefes', methods=['GET'])
def api_get_jefes():
    search = (request.args.get('search') or '').strip()
    id_regional = request.args.get('id_regional', type=int)
    id_ciudad = request.args.get('id_ciudad', type=int)
    limit = request.args.get('limit', type=int) or 50
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_create_jefe: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)
    # Ensure the many-to-many table exists (safe to call repeatedly)
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS jefe_regional (
                id_jefe INT NOT NULL,
                id_regional INT NOT NULL,
                PRIMARY KEY(id_jefe, id_regional)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
    except Exception:
        # ignore creation errors here; table may be managed by migrations
        pass

    # Return jefes with a list of regional ids (supports filtering by regional and optionally by city)
    # Behavior:
    # - if only id_regional provided: include jefes whose legacy column matches OR have an assignment in jefe_regional
    # - if id_regional AND id_ciudad provided: additionally require the jefe to have at least one usuario OR one linea in that city
    sql = """
        SELECT j.id_jefe, j.nombre_jefe, j.id_regional AS legacy_reg, GROUP_CONCAT(DISTINCT jr.id_regional) AS regionals
        FROM jefes j
        LEFT JOIN jefe_regional jr ON jr.id_jefe = j.id_jefe
    """
    where = []
    params = []
    if search:
        where.append("LOWER(j.nombre_jefe) LIKE %s")
        params.append(f"%{search.lower()}%")

    # Regional filter (legacy OR jefe_regional)
    if id_regional:
        where.append("(j.id_regional = %s OR EXISTS (SELECT 1 FROM jefe_regional x WHERE x.id_jefe = j.id_jefe AND x.id_regional = %s))")
        params.extend([id_regional, id_regional])

    # City filter: only when id_ciudad provided, require the jefe to have operational presence in that city
    if id_ciudad:
        # require at least one usuario in that city OR at least one linea in that city assigned to this jefe
        where.append("(EXISTS (SELECT 1 FROM usuarios u WHERE u.id_jefe = j.id_jefe AND u.id_ciudad = %s) OR EXISTS (SELECT 1 FROM lineas l WHERE l.id_jefe = j.id_jefe AND l.id_ciudad = %s))")
        params.extend([id_ciudad, id_ciudad])

    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " GROUP BY j.id_jefe ORDER BY j.nombre_jefe LIMIT %s"
    params.append(limit)

    cur.execute(sql, params)
    rows = cur.fetchall()
    result = []
    for r in rows:
        regional_ids = []
        if r.get('regionals'):
            regional_ids = [int(x) for x in str(r['regionals']).split(',') if x]
        if r.get('legacy_reg'):
            try:
                lr = int(r.get('legacy_reg'))
                if lr not in regional_ids:
                    regional_ids.append(lr)
            except Exception:
                pass
        result.append({"id_jefe": r['id_jefe'], "nombre_jefe": r['nombre_jefe'], "regional_ids": regional_ids})
    cur.close()
    db.close()
    return jsonify(result)


@app.route('/api/jefes', methods=['POST'])
def api_create_jefe():
    data = request.get_json(force=True)
    nombre = (data.get('nombre_jefe') or '').strip()
    regional_ids = data.get('regional_ids') or []
    # sanitize regional_ids: remove empty values and coerce to int
    try:
        if isinstance(regional_ids, list):
            regional_ids = [int(x) for x in regional_ids if x not in (None, '', 'null')]
        else:
            # if client sent a single value, try to coerce
            regional_ids = [int(regional_ids)] if regional_ids else []
    except Exception as e:
        app.logger.exception('Invalid regional_ids payload')
        return make_response(jsonify({"error": "regional_ids inválidos."}), 400)
    # Backwards compatibility: accept single id_regional
    if not regional_ids and data.get('id_regional'):
        regional_ids = [int(data.get('id_regional'))]
    if not nombre or not regional_ids:
        return make_response(jsonify({"error": "Nombre y al menos una regional obligatorios."}), 400)
    db = conectar_db()
    cur = db.cursor(dictionary=True)
    # opcional unicidad por nombre
    cur.execute("SELECT COUNT(1) AS cnt FROM jefes WHERE nombre_jefe = %s", (nombre,))
    if cur.fetchone()['cnt'] > 0:
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Jefe ya existe."}), 409)

    try:
        # ensure join table exists
        try:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS jefe_regional (
                    id_jefe INT NOT NULL,
                    id_regional INT NOT NULL,
                    PRIMARY KEY(id_jefe, id_regional)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception:
            pass

        # For compatibility with existing schema that may have an id_regional NOT NULL column,
        # detect whether the column exists and insert accordingly.
        try:
            db_name = getattr(db, 'database', None) or 'gestion_lineas'
            cur.execute("SELECT COUNT(1) AS cnt FROM information_schema.columns WHERE table_schema = %s AND table_name = 'jefes' AND column_name = 'id_regional'", (db_name,))
            has_legacy = cur.fetchone().get('cnt', 0) > 0
        except Exception:
            has_legacy = False

        legacy_reg = int(regional_ids[0]) if regional_ids else None
        if has_legacy and legacy_reg is not None:
            cur.execute("INSERT INTO jefes (nombre_jefe, id_regional) VALUES (%s, %s)", (nombre, legacy_reg))
        else:
            cur.execute("INSERT INTO jefes (nombre_jefe) VALUES (%s)", (nombre,))
        # Obtain the last insert id using connection-safe LAST_INSERT_ID()
        try:
            cur.execute("SELECT LAST_INSERT_ID() AS id")
            new_id = cur.fetchone().get('id')
        except Exception:
            # fallback to cursor.lastrowid
            new_id = getattr(cur, 'lastrowid', None)
        # insert relations
        for rid in regional_ids:
            cur.execute("INSERT IGNORE INTO jefe_regional (id_jefe, id_regional) VALUES (%s, %s)", (new_id, int(rid)))
        db.commit()
    except mysql.connector.Error:
        db.rollback()
        cur.close()
        db.close()
        app.logger.exception('MySQL error creating jefe')
        return make_response(jsonify({"error": "Error interno de base de datos."}), 500)
    except Exception:
        db.rollback()
        cur.close()
        db.close()
        app.logger.exception('Unexpected error creating jefe')
        return make_response(jsonify({"error": "Error interno."}), 500)
    cur.execute("SELECT id_jefe, nombre_jefe FROM jefes WHERE id_jefe = %s", (new_id,))
    r = cur.fetchone()
    cur.close()
    db.close()
    return make_response(jsonify({"id_jefe": r['id_jefe'], "nombre_jefe": r['nombre_jefe'], "regional_ids": regional_ids}), 201)


@app.route('/api/jefes/<int:id_jefe>', methods=['GET'])
def api_get_jefe(id_jefe):
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_get_jefe: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)
    cur.execute("SELECT id_jefe, nombre_jefe, id_regional FROM jefes WHERE id_jefe = %s", (id_jefe,))
    row = cur.fetchone()
    if not row:
        cur.close(); db.close()
        return make_response(jsonify({"error": "Jefe no existe."}), 404)
    # collect regionals from join table
    cur.execute("SELECT GROUP_CONCAT(id_regional) AS regs FROM jefe_regional WHERE id_jefe = %s", (id_jefe,))
    jr = cur.fetchone()
    regional_ids = []
    if jr and jr.get('regs'):
        try:
            regional_ids = [int(x) for x in str(jr['regs']).split(',') if x]
        except Exception:
            regional_ids = []
    # include legacy column if present
    if row.get('id_regional'):
        try:
            lr = int(row.get('id_regional'))
            if lr not in regional_ids:
                regional_ids.append(lr)
        except Exception:
            pass
    cur.close(); db.close()
    return jsonify({"id_jefe": row.get('id_jefe'), "nombre_jefe": row.get('nombre_jefe'), "regional_ids": regional_ids})


@app.route('/api/jefes/<int:id_jefe>', methods=['PUT'])
def api_update_jefe(id_jefe):
    data = request.get_json(force=True)
    nombre = (data.get('nombre_jefe') or '').strip()
    regional_ids = data.get('regional_ids') or []
    try:
        if isinstance(regional_ids, list):
            regional_ids = [int(x) for x in regional_ids if x not in (None, '', 'null')]
        else:
            regional_ids = [int(regional_ids)] if regional_ids else []
    except Exception:
        return make_response(jsonify({"error": "regional_ids inválidos."}), 400)
    if not nombre or not regional_ids:
        return make_response(jsonify({"error": "Nombre y al menos una regional obligatorios."}), 400)

    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_update_jefe: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # verify exists
    cur.execute("SELECT COUNT(1) AS cnt FROM jefes WHERE id_jefe = %s", (id_jefe,))
    if cur.fetchone().get('cnt', 0) == 0:
        cur.close(); db.close()
        return make_response(jsonify({"error": "Jefe no existe."}), 404)

    try:
        # update name
        cur.execute("UPDATE jefes SET nombre_jefe = %s WHERE id_jefe = %s", (nombre, id_jefe))
        # update legacy column if present
        try:
            db_name = getattr(db, 'database', None) or 'gestion_lineas'
            cur.execute("SELECT COUNT(1) AS cnt FROM information_schema.columns WHERE table_schema = %s AND table_name = 'jefes' AND column_name = 'id_regional'", (db_name,))
            has_legacy = cur.fetchone().get('cnt', 0) > 0
        except Exception:
            has_legacy = False
        if has_legacy:
            # set legacy to first regional id
            legacy_val = regional_ids[0] if regional_ids else None
            cur.execute("UPDATE jefes SET id_regional = %s WHERE id_jefe = %s", (legacy_val, id_jefe))

        # Replace entries in jefe_regional
        cur.execute("DELETE FROM jefe_regional WHERE id_jefe = %s", (id_jefe,))
        for rid in regional_ids:
            cur.execute("INSERT IGNORE INTO jefe_regional (id_jefe, id_regional) VALUES (%s, %s)", (id_jefe, int(rid)))
        db.commit()
    except mysql.connector.Error:
        db.rollback(); cur.close(); db.close(); app.logger.exception('MySQL error updating jefe');
        return make_response(jsonify({"error": "Error interno de base de datos."}), 500)
    except Exception:
        db.rollback(); cur.close(); db.close(); app.logger.exception('Unexpected error updating jefe');
        return make_response(jsonify({"error": "Error interno."}), 500)

    # return updated resource
    cur.execute("SELECT id_jefe, nombre_jefe FROM jefes WHERE id_jefe = %s", (id_jefe,))
    row = cur.fetchone()
    # fetch regionals
    cur.execute("SELECT GROUP_CONCAT(id_regional) AS regs FROM jefe_regional WHERE id_jefe = %s", (id_jefe,))
    jr = cur.fetchone()
    rids = []
    if jr and jr.get('regs'):
        rids = [int(x) for x in str(jr['regs']).split(',') if x]
    cur.close(); db.close()
    return jsonify({"id_jefe": row.get('id_jefe'), "nombre_jefe": row.get('nombre_jefe'), "regional_ids": rids})


# =========================
# API: Ciudades (GET, POST)
# =========================
@app.route('/api/ciudades', methods=['GET'])
def api_get_ciudades():
    """Listado de ciudades, opcionalmente filtrado por regional o búsqueda"""
    search = (request.args.get('search') or '').strip()
    id_regional = request.args.get('id_regional', type=int)
    limit = request.args.get('limit', type=int) or 200
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_get_ciudades: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)
    sql = "SELECT id_ciudad, nombre_ciudad, id_regional FROM ciudades"
    where = []
    params = []
    if search:
        where.append("LOWER(nombre_ciudad) LIKE %s")
        params.append(f"%{search.lower()}%")
    if id_regional:
        where.append("id_regional = %s")
        params.append(id_regional)
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY nombre_ciudad LIMIT %s"
    params.append(limit)
    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    db.close()
    return jsonify([{"id_ciudad": r['id_ciudad'], "nombre_ciudad": r['nombre_ciudad']} for r in rows])


@app.route('/api/regionales', methods=['POST'])
def api_create_regional():
    data = request.get_json(force=True)
    nombre = (data.get('nombre_regional') or '').strip()
    if not nombre:
        return make_response(jsonify({"error": "Nombre de regional obligatorio."}), 400)
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_create_regional: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # unicidad por nombre
    cur.execute("SELECT COUNT(1) AS cnt FROM regionales WHERE nombre_regional = %s", (nombre,))
    if cur.fetchone()['cnt'] > 0:
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Regional ya existe."}), 409)

    try:
        cur.execute("INSERT INTO regionales (nombre_regional) VALUES (%s)", (nombre,))
        new_id = cur.lastrowid
        db.commit()
    except mysql.connector.Error as e:
        print(f"api_create_regional: DB insert error: {e}")
        db.rollback()
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Error interno al crear regional."}), 500)

    cur.execute("SELECT id_regional, nombre_regional FROM regionales WHERE id_regional = %s", (new_id,))
    r = cur.fetchone()
    cur.close()
    db.close()
    return make_response(jsonify({"id_regional": r['id_regional'], "nombre_regional": r['nombre_regional']}), 201)


@app.route('/api/ciudades', methods=['POST'])
def api_create_ciudad():
    data = request.get_json(force=True)
    nombre = (data.get('nombre_ciudad') or '').strip()
    id_regional = data.get('id_regional')
    if not nombre or not id_regional:
        return make_response(jsonify({"error": "Nombre de ciudad y regional obligatorios."}), 400)
    try:
        db = conectar_db()
        cur = db.cursor(dictionary=True)
    except Exception as e:
        print(f"api_create_ciudad: DB connection failed: {e}")
        return make_response(jsonify({"error": "Error connecting to database."}), 503)

    # verificar regional existe
    cur.execute("SELECT COUNT(1) AS cnt FROM regionales WHERE id_regional = %s", (id_regional,))
    if cur.fetchone()['cnt'] == 0:
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Regional no existe."}), 400)

    # unicidad por nombre dentro de la regional
    cur.execute("SELECT COUNT(1) AS cnt FROM ciudades WHERE nombre_ciudad = %s AND id_regional = %s", (nombre, id_regional))
    if cur.fetchone()['cnt'] > 0:
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Ciudad ya existe en esta regional."}), 409)

    try:
        cur.execute("INSERT INTO ciudades (nombre_ciudad, id_regional) VALUES (%s, %s)", (nombre, id_regional))
        new_id = cur.lastrowid
        db.commit()
    except mysql.connector.Error as e:
        print(f"api_create_ciudad: DB insert error: {e}")
        db.rollback()
        # Handle duplicate key gracefully: return existing ciudad if it exists
        if getattr(e, 'errno', None) == 1062:
            try:
                cur.execute("SELECT id_ciudad, nombre_ciudad FROM ciudades WHERE nombre_ciudad = %s LIMIT 1", (nombre,))
                existing = cur.fetchone()
                if existing:
                    cur.close()
                    db.close()
                    return make_response(jsonify({"error": "Ciudad ya existe.", "id_ciudad": existing['id_ciudad'], "nombre_ciudad": existing['nombre_ciudad']}), 409)
            except Exception:
                pass
        cur.close()
        db.close()
        return make_response(jsonify({"error": "Error interno al crear ciudad."}), 500)

    cur.execute("SELECT id_ciudad, nombre_ciudad FROM ciudades WHERE id_ciudad = %s", (new_id,))
    r = cur.fetchone()
    cur.close()
    db.close()
    return make_response(jsonify({"id_ciudad": r['id_ciudad'], "nombre_ciudad": r['nombre_ciudad']}), 201)


if __name__ == '__main__':
    app.run(debug=True)
